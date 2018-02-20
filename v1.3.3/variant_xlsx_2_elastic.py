#!/usr/bin/python2.7


import json
import openpyxl
import pprint
import re
import datetime
import random
from openpyxl.utils import get_column_letter 
from collections import defaultdict
from elasticsearch import Elasticsearch

def removew(d):
    
    return  {re.sub(" ","_",str(k)):removew(v) if isinstance(v, dict) else v for k, v in d.iteritems()}

def remove_none(d):
    
    return dict((k, v) for k, v in d.iteritems() if v!=None)


pp = pprint.PrettyPrinter(indent=4)

input_spreadsheet = "/var/www/TPL/ftp/targetedNGSPipeline/MD/HOLD/MP17-579/MP17-579.unfiltered.xlsx"
case_notes = "/var/www/TPL/ftp/targetedNGSPipeline/MD/HOLD/MP17-579/notes.txt"

wb = openpyxl.load_workbook(filename=input_spreadsheet, data_only=True)

es = Elasticsearch()

master_dict = defaultdict(dict)

def create_variant_dict():
    
    def hgvsp_to_hgvsp_short(variant_dict):
        """Takes an hgsvp string with unabbreviation amino acids
           and converts it into the appropriate abbreviation (e.g. Gly12Asp --> G12D.
        """

        # set up amino acid abbreviation dictionary
        amino_acid_abbreviations = {
                                    'Ala' : 'A',
                                    'Arg' : 'R',
                                    'Asn' : 'N',
                                    'Asp' : 'D',
                                    'Cys' : 'C',
                                    'Glu' : 'E',
                                    'Gln' : 'Q',
                                    'Gly' : 'G',
                                    'His' : 'H',
                                    'Ile' : 'I',
                                    'Leu' : 'L',
                                    'Lys' : 'K',
                                    'Met' : 'M',
                                    'Phe' : 'F',
                                    'Pro' : 'P',
                                    'Ser' : 'S',
                                    'Thr' : 'T',
                                    'Trp' : 'W',
                                    'Tyr' : 'Y',
                                    'Val' : 'V',
                                    'Ter' : '*'
                                    }
        
        
        if 'hgvsp' in variant_dict.keys() and 'amino_acids' in variant_dict.keys():
            # capture only the notation part - we don't need the protein identifier
            hgvsp_notation_only = variant_dict['hgvsp'].split(".")[-1]
            # split on 'Ter' for fs variants - otherwise this will sub the 'Ter' in the frameshift notation
            hgvsp_notation_only = hgvsp_notation_only.split("Ter", 1)[0]
            hgvsp_short = hgvsp_notation_only
            for amino_acid, abbreviation in amino_acid_abbreviations.iteritems():
                if re.search(amino_acid, hgvsp_notation_only):
                    hgvsp_short = re.sub(amino_acid, abbreviation, hgvsp_short)

            variant_dict['hgvsp_short'] = hgvsp_short
        
        return variant_dict
            
    
    ws = wb['summary']
    
    variants = []
    headers = []
    row_counter = 0
    for row in ws.rows:
        values = []
        row_counter+=1
        if row_counter==1:
            for cell in row:
                cell.value = re.sub("\.", "", cell.value)
                cell.value = re.sub("#", "", cell.value)
                headers.append(cell.value.lower())
        else:
            for cell in row:
                if re.search(",", str(cell.value)):
                    values.append(cell.value.split(","))
                else:
                    if type(cell.value) is float:
                        values.append(round(cell.value, 3))
                    else:
                        values.append(cell.value)
                
        if len(values)==len(headers):
            # zip headers and values
            # also remove whitespace from headers
            print values
            header_value_dict = removew(dict(zip(headers, values)))
            header_value_dict = remove_none(header_value_dict)
            header_value_dict = defaultdict(lambda: None, header_value_dict)
            print header_value_dict['n_vaf']
            print header_value_dict
            print header_value_dict['comment']
            if 'comment' in header_value_dict.keys():
                if re.search("OK:", str(header_value_dict['comment'])):
                    header_value_dict = hgvsp_to_hgvsp_short(header_value_dict)
                    variants.append(header_value_dict)
    
    
    return variants

def create_cnv_dict():
    
    ws = wb['cnv']
    
    cnvs = []
    headers = []
    row_counter = 0
    for row in ws.rows:
        values = []
        row_counter+=1
        if row_counter==1:
            for cell in row:
                cell.value = re.sub("\#", "", cell.value)
                headers.append(cell.value.lower())
        else:
            for cell in row:

                if re.search(",", str(cell.value)):
                    values.append(cell.value.split(","))
                else:
                    if type(cell.value) is float:
                        values.append(round(cell.value, 3))
                    else:
                        values.append(cell.value)
                
        if len(values)==len(headers):
            # zip headers and values
            # also remove whitespace from headers
            header_value_dict = removew(dict(zip(headers, values)))
            header_value_dict = remove_none(header_value_dict)
            header_value_dict = defaultdict(lambda: None, header_value_dict)
            if re.search("OK:", str(header_value_dict['comment'])):
                cnvs.append(header_value_dict)
    
    return cnvs

def create_fusion_dict():
    
    ws = wb['fusion']
    
    fusions = []
    headers = []
    row_counter = 0
    for row in ws.rows:
        values = []
        row_counter+=1
        if row_counter==1:
            for cell in row:
                if str(cell.value) == "3'/5' Imbalance":
                    headers.append('imbalance_assay')
                elif str(cell.value) == "Genes(Exons)":
                    headers.append('genes_exons')
                elif str(cell.value) == "COSMIC/NCBI":
                    headers.append('COSMIC-NCBI_id')
                else:
                    headers.append(str(cell.value).lower()) 
            print headers
        else:
            for cell in row:
                values.append(cell.value)
                
        if len(values)==len(headers):
            # zip headers and values
            # also remove whitespace from headers
            header_value_dict = removew(dict(zip(headers, values)))
            header_value_dict = remove_none(header_value_dict)
            header_value_dict = defaultdict(lambda: None, header_value_dict)
            if re.search("OK:", str(header_value_dict['comment'])):
                fusions.append(header_value_dict)
    
    return fusions

def create_specimen_dict():
    
    ws = wb['Specimen']

    specimen = {}
    headers = []
    row_counter = 0
    for row in ws.rows:
        row_counter+=1
        for cell in row:
            if ws['A%s' % row_counter].value == 'Copath#':
                specimen['copath_id'] = str(ws['B%s' % row_counter].value)
            if ws['A%s' % row_counter].value == 'Sequencing done':
                specimen[ws['A%s' % row_counter].value.lower()] = datetime.datetime.strftime(datetime.datetime.date(ws['B%s' % row_counter].value), "%Y-%m-%d")
            else:
                specimen[ws['A%s' % row_counter].value.lower()] = str(ws['B%s' % row_counter].value)

    specimen.pop('copath#')

    specimen = removew(specimen)
    specimen = remove_none(specimen)
    
    try:
        specimen['notes'] = open(case_notes).read()
    except:
        pass

    return specimen
                

master_dict['variant'] = create_variant_dict()
master_dict['cnv'] = create_cnv_dict()
master_dict['fusion'] = create_fusion_dict()
master_dict['specimen'] = create_specimen_dict()

entry = json.loads(json.dumps(master_dict))

internal_case_id = input_spreadsheet.split("/")[-1].strip(".xlsx")
internal_analysis_id = internal_case_id + "_" + datetime.datetime.strftime(datetime.datetime.today(), "%m-%d-%Y") + "_" + ''.join(random.choice('0123456789ABCDEF') for i in range(16))

for index_type in master_dict.keys():
    if type(master_dict[index_type]) is not dict:
        if type(master_dict[index_type]) is list:
            for item in master_dict[index_type]:
                item['internal_case_id'] = internal_case_id
                item['internal_analysis_id'] = internal_analysis_id
                entry = json.loads(json.dumps(item))
                pp.pprint(entry)
                #res = es.create(index='dna-seq', doc_type=index_type, body=entry)
    else:
        master_dict[index_type]['internal_case_id'] = internal_case_id
        master_dict[index_type]['internal_analysis_id'] = internal_analysis_id
        item = master_dict[index_type]
        entry = json.loads(json.dumps(item))
        print entry
        #res = es.create(index='dna-seq', doc_type=index_type, body=entry)

#res = es.create(index='variants', doc_type='variants', body=entry, id=str(master_dict['id'])+str('_')+str(datetime.datetime.today().date()))
#res = es.create(index='dna-seq', doc_type='variants', body=entry, id=str(master_dict['id']))
#es.indices.refresh(index="dna-seq")
          
            


