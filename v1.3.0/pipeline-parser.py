#!/usr/bin/python2.7

import os
import re
import pprint
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import BLUE
from openpyxl.utils import get_column_letter
from collections import defaultdict
from collections import OrderedDict
from natsort import natsorted
import argparse

def main():
    vers="1.0"
    desc="""This tool is designed to parse pipeline output files and convert into a more readable format (e.g. a spreadsheet)"""
    parser = argparse.ArgumentParser(description=desc, version=vers)
    
    general_arguments = parser.add_argument_group('GENERAL OPTIONS')
    general_arguments.add_argument("--mode","-m", help="what pipeline will be parsed",dest='mode',action='store',nargs=1, choices=['tumor-normal'])
    
    tumor_normal_mode = parser.add_argument_group('TUMOR-NORMAL PIPELINE MODE')
    tumor_normal_mode.add_argument("--autodetect_tn_pipeline_files","-a", help="Autodetect tumor-normal pipeline output files in current working directory",dest='autodetect',action='store_true')
    tumor_normal_mode.add_argument("--variant-json", help="One or more ensembl.vep.json files.  Must follow specific naming according to pipeline specifications",dest='variant_jsons',action='append',nargs="+", default=None)
    tumor_normal_mode.add_argument("--ionreporter-cnv-vcf", help="IonReporter CNV .vcf",dest='ir_cnv_vcf',action='store',nargs=1)
    tumor_normal_mode.add_argument("--ionreporter-cnv-tsv", help="IonReporter CNV .tsv",dest='ir_cnv_tsv',action='store',nargs=1)
    tumor_normal_mode.add_argument("--ionreporter-fusion-vcf", help="IonReporter Fusion .vcf",dest='ir_fusion_vcf',action='store',nargs=1,default=None)
    tumor_normal_mode.add_argument("--specimen-json", help="Specimen .json",dest='specimen_json',action='store',nargs=1,default=None)
    tumor_normal_mode.add_argument("--output-basename", help="Output basename",dest='output_basename',action='store',nargs=1,default="output")

    
    args = parser.parse_args()

    print args.ir_cnv_tsv
    print args.variant_jsons
    
    class CNV_TSV_Fields:
        def __init__(self, split_tsv_line):
            self.chr = "chr"+split_tsv_line[0]
            self.pos = split_tsv_line[1]
            self.type = split_tsv_line[2]
            self.length = split_tsv_line[4]
            self.iscn = split_tsv_line[11]
            match = re.search("x(\d+?)", self.iscn)
            self.ploidy = match.group(1)
            
            self.confidence = split_tsv_line[12]
            self.precision = split_tsv_line[13]
            self.genes = split_tsv_line[17]
    
    class CNV_VCF_Fields:
        def __init__(self, split_vcf_line, vcf_line):
            self.chr = split_vcf_line[0]
            self.pos = split_vcf_line[1]
            self.end = re.search("END=(\d+?);", vcf_line).group(1)
            self.numtiles = re.search("NUMTILES=(\d+?);", vcf_line).group(1)
    
    class INFO_VCF_Fields:
        def __init__(self, info_dict, filename):
            if re.search("ionreporter", filename):
                if 'FDP' in info_dict.keys():
                    self.depth = info_dict['FDP']
                else:
                    try:
                        if 'DP' in info_dict.keys():
                            self.depth = info_dict['DP']
                    except:
                        self.depth = info_dict['FDP']
            
    class FORMAT_VCF_Fields:
        def __init__(self, format_dict, filename, alt):
            if re.search("ionreporter", filename):
                if 'FDP' in format_dict.keys():
                    self.depth = format_dict['FDP']
                else:
                    try:
                        if 'DP' in format_dict.keys():
                            self.depth = format_dict['DP']
                    except:
                        self.depth = format_dict['FDP']
                try:
                    self.fro = int(format_dict['FRO'])
                    self.fao = int(format_dict['FAO'])
                    self.vaf = float(format_dict['FAO']) / float(int(format_dict['FAO']) + int(format_dict['FRO']))
                except Exception, e:
                    print "ERROR: Unable to define format fields"
                    print str(e)
                    self.fro, self.fao, self.vaf = ("UNK" for i in range(3))
                try:
                    self.fsrb = float(float(format_dict['FSAF']) / float(int(format_dict['FSAF']) + int(format_dict['FSAR'])))
                    self.rsrb = float(float(format_dict['FSAR']) / float(int(format_dict['FSAF']) + int(format_dict['FSAR'])))
                except Exception, e:
                    print "ERROR10:10 Unable to define format fields"
                    print str(e)
                    self.fsrb, self.rsrb = ("UNK" for i in range(2))
            elif re.search("mutect2", filename):
                try:
                    split_AD = format_dict['AD'].split(",")
                    self.fro = split_AD[0]
                    self.fao = split_AD[1]
                    self.depth = int(self.fao) + int(self.fro)
                    self.vaf = format_dict['AF']
                    self.fsrb = "UNK"
                    self.rsrb = "UNK"
                except Exception, e:
                    print "ERROR: Unable to define format fields"
                    print str(e)
                    self.depth, self.fsrb, self.rsrb, self.vaf, self.fro, self.fao = ("UNK" for i in range(6)) 
            elif re.search("strelka", filename):
                # Check if INDEL, process accordingly
                # TIR tag is only present for INDELs
                try:
                    self.depth = format_dict['DP']
                    if 'TIR' in format_dict.keys():
                        self.fao = format_dict['TIR'].split(",")[0]
                        self.fro = int(self.depth) - int(self.fao)
                        self.vaf = float(float(self.fao) / float(self.depth))
                        self.fsrb = "UNK"
                        self.rsrb = "UNK"
                    else:
                        self.fao = format_dict["%sU" % alt].split(",")[0]
                        self.fro = int(self.depth) - int(self.fao)
                        self.vaf = float(float(self.fao) / float(self.depth))
                        self.fsrb = "UNK"
                        self.rsrb = "UNK"
                except Exception, e:
                    print "ERROR: Unable to define format fields"
                    print str(e) 
                    self.depth, self.fsrb, self.rsrb, self.vaf, self.fro, self.fao = ("UNK" for i in range(6)) 
                
    class VCF_Fields():
        def __init__(self,vcf_input, program, status):
            
            split_input = vcf_input.strip().split("\t")
            reconstructed_file_basename = program + "." + status + ".json"
            
            try:
                self.chrom = split_input[0]
            except:
                self.chrom = ""
            try:
                self.pos = split_input[1]
            except:
                self.pos = ""
            try:
                self.id = split_input[2]
            except:
                self.id = ""
            try:
                self.ref = split_input[3]
            except:
                self.ref = ""
            try:
                self.alt = split_input[4]
            except:
                self.alt = ""
            try:
                self.qual = split_input[5]
            except:
                self.qual = ""
            try:
                self.filter = split_input[6]
            except:
                self.filter = ""
            try:
                self.info = split_input[7]
                self.split_info = self.info.split(";")
                self.info_dict = defaultdict(int)
                for k_v_pair in self.split_info:
                    k_v_pair_split = k_v_pair.split("=")
                    if len(k_v_pair_split) > 1 and len(k_v_pair_split) == 2:
                        k, v = k_v_pair_split
                        
                        self.info_dict[k] = v
                
                self.info_fields = INFO_VCF_Fields(self.info_dict, reconstructed_file_basename)
            except Exception, e:
                print str(e)
                self.info = ""
            
            
            # FORMAT  
            try:
                self.format = split_input[8]
                self.format_keys = self.format.split(":")
    
                if program == "strelka":
                    tumor_order = split_input[10]
                    normal_order = split_input[9]        
                else:
                    tumor_order = split_input[9]
                    try:
                        normal_order = split_input[10]
                    except:
                        normal_order = ["UNK" for i in range(len(self.format_keys))]
                try:
                    self.format_tumor_values = tumor_order.split(":")
                    tmp_dict = dict(zip(self.format_keys, self.format_tumor_values))
                    self.format_tumor_dict = defaultdict(lambda: "UNK", tmp_dict)
                    self.format_tumor = FORMAT_VCF_Fields(self.format_tumor_dict, reconstructed_file_basename, self.alt)
                except Exception, e:
                    self.format_tumor_dict = defaultdict(lambda: "UNK")
                    self.format_tumor = FORMAT_VCF_Fields(self.format_tumor_dict, reconstructed_file_basename, self.alt)
                try:
                    self.format_normal_values = normal_order.split(":")
                    tmp_dict = dict(zip(self.format_keys, self.format_normal_values))
                    self.format_normal_dict = defaultdict(lambda: "UNK", tmp_dict)
                    self.format_normal = FORMAT_VCF_Fields(self.format_normal_dict, reconstructed_file_basename, self.alt)
                except Exception, e:
                    self.format_normal_dict = defaultdict(lambda: "UNK")
                    self.format_normal = FORMAT_VCF_Fields(self.format_normal_dict, reconstructed_file_basename, self.alt)
                
            except:
                self.format = ""
    
    class FusionVCFEntry:
        def __init__(self, line):
    
            class ImbalanceAssay:
                def __init__(self, split_line, line):
                    self.variant_id = split_line[2]
                    self.locus = split_line[0] + ":" + str(split_line[1]) + " - " + split_line[0] + ":" + re.search("THREE_PRIME_POS=(\d+?);", line).group(1)
                    self.genes_and_exons = re.search("(\w+?)\.\w+?\..+?", self.variant_id).group(1)
                    self.read_count = re.search("READ_COUNT=(.+?);", line).group(1)
                    self.oncomine_gene_class = ""
                    self.oncomine_variant_class = ""
                    self.detection = "See Doc."
                    self.imbalance_assay = re.search("5P_3P_ASSAYS=(.+?);", line).group(1)
                    self.annotation = ""
                    self.rpm = re.search("RPM=(.+?);", line).group(1)
                    
            class ExprControl:
                def __init__(self, split_line, line):
                    self.variant_id = split_line[2]
                    self.locus = split_line[0] + ":" + str(split_line[1])
                    self.genes_and_exons = re.search("(\w+?)\.\w+?\..+?", self.variant_id).group(1)
                    self.read_count = re.search("READ_COUNT=(\d+);", line).group(1)
                    self.oncomine_gene_class = ""
                    self.oncomine_variant_class = ""
                    self.detection = "See Doc."
                    self.imbalance_assay = "" 
                    self.annotation = ""
                    self.rpm = re.search("RPM=(.+?);", line).group(1)
    
            class FusionEntry:
                def __init__(self, split_line, line):
                    self.variant_id = re.search("(.+?)_\d", split_line[2]).group(1)
                    self.locus = split_line[0] + ":" + str(split_line[1])
                    self.donor_gene = re.search("(.+?)-(.+?)\.", split_line[2]).group(1)
                    self.partner_gene = re.search("(.+?)-(.+?)\.", split_line[2]).group(2)
                    self.donor_exon = re.search("\.\w+?(\d+)\w+?(\d+).", split_line[2]).group(1)
                    self.parter_exon = re.search("\.\w+?(\d+)\w+?(\d+).", split_line[2]).group(2)
                    self.genes_and_exons = "%s(%s) - %s(%s)" % (self.donor_gene,
                                                                self.donor_exon,
                                                                self.partner_gene,
                                                                self.parter_exon)
                    self.read_count = re.search("READ_COUNT=(\d+?);", line).group(1)
                    try:
                        self.oncomine_gene_class = re.search("\'oncomineGeneClass\':\'(.+?)\'", line).group(1)
                    except:
                        self.oncomine_gene_class = ""
                    try:
                        self.oncomine_variant_class = re.search("\'oncomineVariantClass\':\'(.+?)\'", line).group(1)
                    except:
                        self.oncomine_variant_class = ""
                    
                    if int(self.read_count) > 0:
                        self.detection = "Present"
                    else:
                        self.detection = "Absent"
                    
                    self.imbalance_assay = ""
                    
                    try:
                        self.annotation = re.search("ANNOTATION=(.+?);", line).group(1)
                    except:
                        self.annotation = ""
                    
                    self.rpm = re.search("RPM=(.+?);", line).group(1)
    
            split_line = line.strip().split("\t")
            svtype = re.search("SVTYPE=(\w+?);", line).group(1)
            self.svtype = svtype
            
            if svtype == "Fusion":
                self.fusion = FusionEntry(split_line, line)
            elif svtype == "ExprControl":
                self.expr_control = ExprControl(split_line, line)
            elif svtype == "5p3pAssays":
                self.imbalance_assay = ImbalanceAssay(split_line, line)
    
    def natural_sort(l): 
        convert = lambda text: int(text) if text.isdigit() else text.lower() 
        alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key) ] 
        return sorted(l, key = alphanum_key)
    
    def traverse(o, tree_types=(list, tuple)):
        if isinstance(o, tree_types):
            for value in o:
                for subvalue in traverse(value, tree_types):
                    yield subvalue
        else:
            yield o
    
    def tn_pipeline_output_processing():
    
        def validate_input():
        
            possible_extensions = [".strelka.somatic.json",
                                   ".mutect2.somatic.json",
                                   ".ionreporter.tsv",
                                   ".ionreporter.somatic.json",
                                   ".ionreporter.loh.json",
                                   ".ionreporter.germline.json",
                                   ".ionreporter.fusions.vcf",
                                   ".ionreporter.cnv.vcf",
                                   ".specimen.json"
                                   ]
        
            detected_files = []
        
            for _file in os.listdir(os.getcwd()):
                for possible_extension in possible_extensions:
                    if re.search(possible_extension, _file):
                        
                        detected_files.append(os.path.abspath("%s/%s" % (os.getcwd(), _file)))
            
            return detected_files
    
        def autofit_columns():
            for sheet in wb.get_sheet_names():
                active_sheet = wb.get_sheet_by_name(sheet)
                                
                dims = {}
                for row in active_sheet.rows:
                    for cell in row:
                        if cell.value:
                            if re.search("=HYPERLINK", str(cell.value)):
                                match = re.search("""=HYPERLINK.+?(\".+?\").+?(\".+?\")""", cell.value)
                                displayed_cell_value = match.group(2).strip('"') 
                                dims[cell.column] = max((dims.get(cell.column, 0), len(str(displayed_cell_value))+5))
                            else:
                                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))+3))
                for col, value in dims.items():
                    active_sheet.column_dimensions[col].width = value
    
                if sheet=='summary':
                    active_sheet.freeze_panes = active_sheet.cell("%s%s" % ("G", 2) )
                elif sheet=='transcripts':
                    active_sheet.freeze_panes = active_sheet.cell("%s%s" % ("C", 2) )
                elif sheet=='co-located':
                    active_sheet.freeze_panes = active_sheet.cell("%s%s" % ("C", 2) )
                elif sheet=='regulatory':
                    active_sheet.freeze_panes = active_sheet.cell("%s%s" % ("G", 2) )
                elif sheet=='cnv':
                    active_sheet.freeze_panes = active_sheet.cell("%s%s" % ("C", 2) )
                elif sheet=='fusion':
                    active_sheet.freeze_panes = active_sheet.cell("%s%s" % ("C", 2) )
        
        def determine_best_colocated_variant(variant_entry):
            variant_entry = defaultdict(lambda: None, variant_entry)
            
            best_colocated_variant = {}
            
            if variant_entry['colocated_variants'] is None:
                return None
            else:
                for colocated_variant in variant_entry['colocated_variants']:
                    if re.match("COSM", colocated_variant['id']):
                        best_colocated_variant = colocated_variant
                    else:
                        if not best_colocated_variant:
                            best_colocated_variant = colocated_variant
                        else:
                            pass
    
                return best_colocated_variant
    
        def determine_best_transcript(variant_entry):
            
            best_transcript = {}
            
            
            for consequence in variant_entry['transcript_consequences']:
                
                try:
                    if consequence['canonical'] == 1:
                        if not best_transcript:
                            best_transcript = consequence
                        else:
                            if re.match("ENST", consequence['transcript_id']):
                                best_transcript = consequence
                            else:
                                pass
                except KeyError:
                    # if canonical flag doesn't exist
                    if not best_transcript:
                        best_transcript = consequence
                        best_transcript['canonical'] = 0
                    else:
                        try:
                            if best_transcript['canonical']:
                                pass
                            else:
                                best_transcript = consequence
                                best_transcript['canonical'] = 0
                        except KeyError:
                            best_transcript = consequence
                            best_transcript['canonical'] = 0
    
            return best_transcript
    
        def parse_variant_json(filename):
            """Parse a variant json from ensembl vep."""
            
            def enhance_variant_dict(line_obj):
                """Enhance the ensembl vep json dict with vcf info."""
                
                enhanced_variant_dict = line_obj.json_loaded.copy()
                enhanced_variant_dict['comment'] = line_obj.comment
                if line_obj.status == "loh":
                    enhanced_variant_dict['status'] = line_obj.status.upper()
                else:
                    enhanced_variant_dict['status'] = line_obj.status.capitalize()
                if line_obj.program == "ionreporter":
                    enhanced_variant_dict['program'] = "ir"
                elif line_obj.program == "mutect2":
                    enhanced_variant_dict['program'] = "mt2"
                elif line_obj.program == "strelka":
                    enhanced_variant_dict['program'] = "stlka"
                elif line_obj.program == "varscan":
                    enhanced_variant_dict['program'] = "vs"
                enhanced_variant_dict['vcf_pos'] = line_obj.vcf.pos
                enhanced_variant_dict['ref_allele'] = line_obj.vcf.ref
                enhanced_variant_dict['alt_allele'] = line_obj.vcf.alt
                enhanced_variant_dict['T Cov.'] = line_obj.vcf.format_tumor.depth
                enhanced_variant_dict['T FSRB'] = line_obj.vcf.format_tumor.fsrb
                enhanced_variant_dict['T RSRB'] = line_obj.vcf.format_tumor.rsrb
                enhanced_variant_dict['T VAF'] = line_obj.vcf.format_tumor.vaf
                enhanced_variant_dict['T FRO'] = line_obj.vcf.format_tumor.fro 
                enhanced_variant_dict['T FAO'] = line_obj.vcf.format_tumor.fao
                enhanced_variant_dict['N Cov.'] = line_obj.vcf.format_normal.depth
                enhanced_variant_dict['N FSRB'] = line_obj.vcf.format_normal.fsrb
                enhanced_variant_dict['N RSRB'] = line_obj.vcf.format_normal.rsrb
                enhanced_variant_dict['N VAF'] = line_obj.vcf.format_normal.vaf
                enhanced_variant_dict['N FRO'] = line_obj.vcf.format_normal.fro 
                enhanced_variant_dict['N FAO'] = line_obj.vcf.format_normal.fao
                return enhanced_variant_dict
            
            with open(filename, "r") as f:
                for line in f.readlines():
                
                    class VariantEntry:
                        def __init__(self, line, filename):
    
                            def parse_input_line(json_loaded):
                                try:
                                    vcf_input = json_loaded['input']
                                except KeyError:
                                    "ERROR: 'input' does not exist!"
                                    
                                return vcf_input
                            
                            split_filename = filename.split(".")
                            program = split_filename[-3]
                            status = split_filename[-2]
                            self.program = program
                            self.status = status
                            
                            json_loaded = json.loads(line)
                            self.json_loaded = json_loaded
                            
                            vcf_input = parse_input_line(json_loaded)                                
                            
                            
                            vcf_input_obj = VCF_Fields(vcf_input, program, status)
                            
                            try:
                                self.comment = ""
                            except:
                                self.comment = ""
                                
                            self.vcf = vcf_input_obj
    
                            
                    line_obj = VariantEntry(line, filename)
                    enhanced_variant_dict = enhance_variant_dict(line_obj)
    
                    master_json_list.append(enhanced_variant_dict)
    
        def open_output_xlsx():
            """Open a workbook, and load it into memory."""
            
            wb = openpyxl.Workbook()
            return wb
    
        def create_variant_sheets():
            """Create the following sheets in the workbook."""
        
            summary_sheet = wb.active
            summary_sheet.title = "summary"
            transcripts_sheet = wb.create_sheet(title="transcripts")
            colocated_sheet = wb.create_sheet(title="co-located")
            #cnv_sheet = wb.create_sheet(title="cnv")
            regulatory_sheet = wb.create_sheet(title="regulatory")
            #fusion_sheet = wb.create_sheet(title="fusion")
            specimen_sheet = wb.create_sheet(title="Specimen")
        
        def create_sheet(sheetname):
            
            wb.create_sheet(title=sheetname)
        
        def write_variant_sheets_headers():
            """Write variant spreadsheet headers."""
           
            summary_sheet_headers = ['comment',
                                     '#chr',
                                     'gene',
                                     'pos',
                                     'ref',
                                     'alt',
                                     'source',
                                     'status',
                                     'main consequence',
                                     'T Cov.',
                                     'T FSRB',
                                     'T RSRB',
                                     'T VAF',
                                     'T FRO',
                                     'T FAO',
                                     'N Cov.',
                                     'N FSRB',
                                     'N RSRB',
                                     'N VAF',
                                     'N FRO',
                                     'N FAO',
                                     'transcript link',
                                     'transcript id',
                                     'hgvsc',
                                     'canonical',
                                     'biotype',
                                     'consequence',
                                     'ccds',
                                     'exon',
                                     'protein id',
                                     'hgvsp',
                                     'codons',
                                     'amino acids',
                                     'strand',
                                     'sift',
                                     'polyphen',
                                     'swissprot',
                                     'trembl',
                                     'co-located link',
                                     'co-located id',
                                     'co-located somatic',
                                     'cl-start',
                                     'cl-end',
                                     'cl-strand',
                                     'cl-allele',
                                     'cl-minor allele freq',
                                     'cl-minor allele',
                                     'cl-clin sig',
                                     'cl-pubmed',
                                     'assembly name'
                                     ]
           
            transcript_sheet_headers = ['variant id',
                                        'gene',
                                        'transcript id',
                                        'hgvsc',
                                        'canonical',
                                        'biotype',
                                        'consequence',
                                        'ccds',
                                        'exon',
                                        'protein id',
                                        'hgvsp',
                                        'codons',
                                        'amino acids',
                                        'strand',
                                        'sift',
                                        'polyphen',
                                        'swissprot',
                                        'trembl',
                                        ]
            
            colocated_sheet_headers = ['variant',
                                       'known id',
                                       'somatic',
                                       'start',
                                       'end',
                                       'strand',
                                       'allele',
                                       'minor allele freq',
                                       'minor allele',
                                       'clin sig',
                                       'pubmed'
                                       ]
            
            regulatory_sheet_headers = ['comment',
                                        '#chr',
                                        'closest gene',
                                        'pos',
                                        'ref',
                                        'alt',
                                        'source',
                                        'status',
                                        'impact',
                                        'biotype',
                                        'consequence',
                                        'regulatory feature id',
                                        'T Cov.',
                                        'T VAF',
                                        'T FRO',
                                        'T FAO',
                                        'N Cov.',
                                        'N VAF',
                                        'N FRO',
                                        'N FAO',
                                        'co-located link',
                                        'co-located id',
                                        'co-located somatic',
                                        'cl-start',
                                        'cl-end',
                                        'cl-strand',
                                        'cl-allele',
                                        'cl-minor allele freq',
                                        'cl-minor allele',
                                        'cl-clin sig',
                                        'cl-pubmed',
                                        'assembly name'
                                        ]

    
            header_dict = {'summary' : summary_sheet_headers,
                           'transcripts' : transcript_sheet_headers,
                           'co-located' : colocated_sheet_headers,
                           'regulatory' : regulatory_sheet_headers
                           #'cnv' : cnv_sheet_headers,
                           #'fusion' : fusion_sheet_headers
                           }
            
            for sheetname in header_dict.keys():
                active_sheet = wb.get_sheet_by_name(sheetname)
                row_counter = 1
                column_counter = 1
                for header in header_dict[sheetname]:
                    active_cell = active_sheet.cell(row=row_counter, column=column_counter)
                    active_sheet.column_dimensions
                    active_cell.value = header
                    active_cell.font = openpyxl.styles.Font(size=10, name="Arial")
                    column_counter += 1
        
            return header_dict
        
        def write_variants():
            """Gather, coordinate, and all other subfunctions necessary for processing variants from the ensembl vep."""
    
            def write_colocated_sheet(variant_entry, row_counter):
                """Given a variant entry, write co-located variants to 'co-located' sheet."""
                if variant_entry['colocated_variants'] == "":
                    pass
                else:
                    for colocated_variant in variant_entry['colocated_variants']:
                        colocated_variant = defaultdict(lambda: "", colocated_variant)
                        
                        chr = "chr%s" % variant_entry['seq_region_name']
                        #pos = variant_entry['vcf_pos']
                        pos = variant_entry['start']
                        variant = chr + "_" + str(pos)
                        known_id = colocated_variant['id']
                        if 'somatic' in colocated_variant.keys():
                            colocated_somatic = colocated_variant['somatic']
                        else:
                            if 'id' in colocated_variant.keys():
                                colocated_somatic = 0
                            else:
                                colocated_somatic = colocated_variant['somatic']
                        colocated_start = colocated_variant['start']
                        colocated_end = colocated_variant['end']
                        colocated_strand = colocated_variant['strand']
                        colocated_allele = colocated_variant['allele_string']
                        if 'exac_adj_maf' in colocated_variant.keys():
                            colocated_minor_allele_freq = colocated_variant['exac_adj_maf']
                        elif 'minor_allele_freq' in colocated_variant.keys():
                            colocated_minor_allele_freq = colocated_variant['minor_allele_freq']
                        else:
                            colocated_minor_allele_freq = colocated_variant['minor_allele_freq']
    
                        if 'exac_adj_allele' in colocated_variant.keys():
                            colocated_minor_allele = colocated_variant['exac_adj_allele']
                        elif 'minor_allele' in colocated_variant.keys():
                            colocated_minor_allele = colocated_variant['minor_allele']
                        else:
                            colocated_minor_allele = colocated_variant['minor_allele']
            
                        if 'clin_sig' in colocated_variant.keys():
                            colocated_clinsig = ",".join(colocated_variant['clin_sig'])
                        else:
                            colocated_clinsig = colocated_variant['clin_sig']
            
                        if 'pubmed' in colocated_variant.keys():
                            colocated_pubmed = ",".join((str(i) for i in colocated_variant['pubmed']))
                        else:
                            colocated_pubmed = colocated_variant['pubmed']
    
    
                        values_to_print = [variant,
                                           known_id,
                                           colocated_somatic,
                                           colocated_start,
                                           colocated_end,
                                           colocated_strand,
                                           colocated_allele,
                                           colocated_minor_allele_freq,
                                           colocated_minor_allele,
                                           colocated_clinsig,
                                           colocated_pubmed
                                           ]
                        
                        print_dict = OrderedDict(zip(header_dict['co-located'], values_to_print))
                        
                        # PRINT TO CO-LOCATED SHEET
                        
                        colocated_sheet = wb.get_sheet_by_name('co-located')
                        column_counter = 1
                        for k in print_dict.keys():
                            colocated_sheet.cell(row=row_counter, column=column_counter).value = print_dict[k]
                            colocated_sheet.cell(row=row_counter, column=column_counter).font = global_font
                            
                            # Special handling for select columns
                            if k == "variant":
                                master_coordinate_dict[print_dict['variant']]['co-located.variant'] = colocated_sheet.cell(row=row_counter, column=column_counter).coordinate
                            
                            column_counter += 1
                        row_counter += 1
                
                return row_counter
    
            def write_transcript_sheet(variant_entry, row_counter):
                """Given a variant entry, write transcript consequences to 'transcripts' sheet."""
                
                if variant_entry['transcript_consequences'] == "":
                    pass
                else:
                    for consequence in variant_entry['transcript_consequences']:
                        #pp.pprint(consequence)
                        consequence = defaultdict(lambda: "", consequence)
                        
                        chr = "chr%s" % variant_entry['seq_region_name']
                        #pos = variant_entry['vcf_pos']
                        pos = variant_entry['start']
                        variant_id = chr + "_" + str(pos)
                        gene = consequence['gene_symbol']
                        transcript_id = consequence['transcript_id']
                        hgvsc = consequence['hgvsc']
                        if 'canonical' in consequence.keys():
                            canonical = consequence['canonical']
                        else:
                            canonical = 0
                        biotype = consequence['biotype']
                        consequence_terms = ",".join(consequence['consequence_terms'])
                        ccds = consequence['ccds']
                        exon = consequence['exon']
                        protein_id = consequence['protein_id']
                        hgvsp = consequence['hgvsp']
                        codons = consequence['codons']
                        amino_acids = consequence['amino_acids']
                        strand = consequence['strand']
                        if not consequence['sift_prediction'] == "":
                            sift = "/".join( [ consequence['sift_prediction'], str(consequence['sift_score']) ] )
                        else:
                            sift = ""
                        if not consequence['polyphen_prediction'] == "":
                            polyphen = "/".join( [ consequence['polyphen_prediction'], str(consequence['polyphen_score']) ] )
                        else:
                            polyphen = ""
                        swissprot = consequence['swissprot']
                        trembl = consequence['trembl']
                        
                        
                        values_to_print = [variant_id,
                                           gene,
                                           transcript_id,
                                           hgvsc,
                                           canonical,
                                           biotype,
                                           consequence_terms,
                                           ccds,
                                           exon,
                                           protein_id,
                                           hgvsp,
                                           codons,
                                           amino_acids,
                                           strand,
                                           sift,
                                           polyphen,
                                           swissprot,
                                           trembl
                                           ]
                        
                        print_dict = OrderedDict(zip(header_dict['transcripts'], values_to_print))
                        
                        # PRINT TO TRANSCRIPT SHEET
                        
                        transcript_sheet = wb.get_sheet_by_name('transcripts')
                        column_counter = 1
                        for k in print_dict.keys():
                            transcript_sheet.cell(row=row_counter, column=column_counter).value = print_dict[k]
                            transcript_sheet.cell(row=row_counter, column=column_counter).font = global_font
                            
                            # Special handling for select columns
                            if k == "variant id":
                                master_coordinate_dict[print_dict['variant id']]['transcripts.variant id'] = transcript_sheet.cell(row=row_counter, column=column_counter).coordinate
                            
                            column_counter += 1
                        row_counter += 1
                        
                return row_counter
    
            def write_regulatory_sheet(variant_entry, row_counter):
                """Given a variant entry, extract regulatory consequences and print to 'regulatory' sheet."""
                
                best_transcript = determine_best_transcript(variant_entry)
                best_transcript = defaultdict(lambda: "", best_transcript)
    
                best_colocated_variant = determine_best_colocated_variant(variant_entry)
                if best_colocated_variant is None:
                    best_colocated_variant = defaultdict(lambda: "")
                else:
                    best_colocated_variant = defaultdict(lambda: "", best_colocated_variant)
                
                if variant_entry['regulatory_feature_consequences'] == "":
                    pass
                else:
                    for consequence in variant_entry['regulatory_feature_consequences']:
                        #pp.pprint(consequence)
                        consequence = defaultdict(lambda: "", consequence)
                        
                        assembly_name = variant_entry['assembly_name']
                        comment = ""
                        chr = "chr%s" % variant_entry['seq_region_name']
                        gene = best_transcript['gene_symbol']
                        #pos = variant_entry['vcf_pos']
                        pos = variant_entry['start']
                        variant_id = chr + "_" + str(pos)
                        ref = variant_entry['ref_allele']
                        alt = variant_entry['alt_allele']
                        source = variant_entry['program']
                        status = variant_entry['status']
                        regulatory_feature_id = consequence['regulatory_feature_id']
                        impact = consequence['impact']
                        biotype = consequence['biotype']
                        consequence_terms = ",".join(consequence['consequence_terms'])
                        tumor_depth = variant_entry['T Cov.']
                        tumor_fsrb = variant_entry['T FSRB']
                        tumor_rsrb = variant_entry['T RSRB']
                        tumor_vaf = variant_entry['T VAF']
                        tumor_fro = variant_entry['T FRO']
                        tumor_fao = variant_entry['T FAO']
                        normal_depth = variant_entry['N Cov.']
                        normal_fsrb = variant_entry['N FSRB']
                        normal_rsrb = variant_entry['N RSRB']
                        normal_vaf = variant_entry['N VAF']
                        normal_fro = variant_entry['N FRO']
                        normal_fao = variant_entry['N FAO']
                        
                        colocated_link = chr + "_" + str(pos)
                        colocated_id = best_colocated_variant['id']
                        if 'somatic' in best_colocated_variant.keys():
                            colocated_somatic = best_colocated_variant['somatic']
                        else:
                            if 'id' in best_colocated_variant.keys():
                                colocated_somatic = 0
                            else:
                                colocated_somatic = best_colocated_variant['somatic']
                        colocated_start = best_colocated_variant['start']
                        colocated_end = best_colocated_variant['end']
                        colocated_strand = best_colocated_variant['strand']
                        colocated_allele = best_colocated_variant['allele_string']
                        if 'minor_allele_freq' in best_colocated_variant.keys():
                            colocated_minor_allele_freq = best_colocated_variant['minor_allele_freq']
                        elif 'exac_adj_maf' in best_colocated_variant.keys():
                            colocated_minor_allele_freq = best_colocated_variant['exac_adj_maf']
                        else:
                            colocated_minor_allele_freq = best_colocated_variant['minor_allele_freq']
            
                        if 'minor_allele' in best_colocated_variant.keys():
                            colocated_minor_allele = best_colocated_variant['minor_allele']
                        elif 'exac_adj_allele' in best_colocated_variant.keys():
                            colocated_minor_allele = best_colocated_variant['exac_adj_allele']
                        else:
                            colocated_minor_allele = best_colocated_variant['minor_allele']
            
                        if 'clin_sig' in best_colocated_variant.keys():
                            colocated_clinsig = ",".join(best_colocated_variant['clin_sig'])
                        else:
                            colocated_clinsig = best_colocated_variant['clin_sig']
            
                        if 'pubmed' in best_colocated_variant.keys():
                            colocated_pubmed = ",".join((str(i) for i in best_colocated_variant['pubmed']))
                        else:
                            colocated_pubmed = best_colocated_variant['pubmed']
                        # convert floats to decimal with 3 decimal points
                
                        try:
                            tumor_vaf = "{0:.3f}".format(float(tumor_vaf))
                        except ValueError:
                            pass
                        try:
                            tumor_fsrb = "{0:.3f}".format(float(tumor_fsrb))
                        except ValueError:
                            pass
                        try:
                            tumor_rsrb = "{0:.3f}".format(float(tumor_rsrb))
                        except ValueError:
                            pass
                        try:
                            normal_vaf = "{0:.3f}".format(float(normal_vaf))
                        except ValueError:
                            pass
                        try:
                            normal_fsrb = "{0:.3f}".format(float(normal_fsrb))
                        except ValueError:
                            pass
                        try:
                            normal_rsrb = "{0:.3f}".format(float(normal_rsrb))
                        except ValueError:
                            pass   
    
                        values_to_print = [comment,
                                           chr,
                                           gene,
                                           pos,
                                           ref,
                                           alt,
                                           source,
                                           status,
                                           impact,
                                           biotype,
                                           consequence_terms,
                                           regulatory_feature_id,
                                           tumor_depth,
                                           tumor_vaf,
                                           tumor_fro,
                                           tumor_fao,
                                           normal_depth,
                                           normal_vaf,
                                           normal_fro,
                                           normal_fao,
                                           colocated_link,
                                           colocated_id,
                                           colocated_somatic,
                                           colocated_start,
                                           colocated_end,
                                           colocated_strand,
                                           colocated_allele,
                                           colocated_minor_allele_freq,
                                           colocated_minor_allele,
                                           colocated_clinsig,
                                           colocated_pubmed,
                                           assembly_name
                                           ]
                        
                        print_dict = OrderedDict(zip(header_dict['regulatory'], values_to_print))
                        
                        # PRINT TO TRANSCRIPT SHEET
                        
                        regulatory_sheet = wb.get_sheet_by_name('regulatory')
                        column_counter = 1
                        for k in print_dict.keys():
                            regulatory_sheet.cell(row=row_counter, column=column_counter).value = print_dict[k]
                            regulatory_sheet.cell(row=row_counter, column=column_counter).font = global_font
                            
                            # Special handling for select columns
    #                         if k == "variant id":
    #                             master_coordinate_dict[print_dict['variant id']]['regulatorys.variant id'] = regulatory_sheet.cell(row=row_counter, column=column_counter).coordinate
    #                          
                            column_counter += 1
                        row_counter += 1
                        
                return row_counter
                
            def write_summary_sheet(summary_values_dict, row_counter):
                """Given an entry dict, print values to 'summary' sheet."""
                
                summary_sheet = wb.get_sheet_by_name('summary')
                column_counter = 1
                for k in summary_values_dict.keys():
                    summary_sheet.cell(row=row_counter, column=column_counter).value = summary_values_dict[k]
                    summary_sheet.cell(row=row_counter, column=column_counter).font = global_font
                    
                    # Special handling for select columns
                    if k == "transcript link":
                        master_coordinate_dict[summary_values_dict['transcript link']]['summary.transcript link'] = summary_sheet.cell(row=row_counter, column=column_counter).coordinate
                    elif k == "co-located link":
                        master_coordinate_dict[summary_values_dict['co-located link']]['summary.co-located link'] = summary_sheet.cell(row=row_counter, column=column_counter).coordinate
    
                    column_counter += 1
                row_counter += 1
                return row_counter
    
            def gather_summary_sheet_values(variant_entry):
                """Gather values for 'summary' sheet, given an ensembl vep json entry."""
                
                
                best_transcript = determine_best_transcript(variant_entry)
                best_transcript = defaultdict(lambda: "", best_transcript)
    
                best_colocated_variant = determine_best_colocated_variant(variant_entry)
                if best_colocated_variant is None:
                    best_colocated_variant = defaultdict(lambda: "")
                else:
                    best_colocated_variant = defaultdict(lambda: "", best_colocated_variant)
    
                # "summary" sheet values to print
                assembly_name = variant_entry['assembly_name']
                
                comment = variant_entry['comment']
                chr = "chr%s" % variant_entry['seq_region_name']
                #pos = variant_entry['vcf_pos']
                pos = variant_entry['start']
                gene = best_transcript['gene_symbol']
                ref = variant_entry['ref_allele']
                alt = variant_entry['alt_allele']
                source = variant_entry['program']
                status = variant_entry['status']
                main_consequence = variant_entry['most_severe_consequence']
                tumor_depth = variant_entry['T Cov.']
                tumor_fsrb = variant_entry['T FSRB']
                tumor_rsrb = variant_entry['T RSRB']
                tumor_vaf = variant_entry['T VAF']
                tumor_fro = variant_entry['T FRO']
                tumor_fao = variant_entry['T FAO']
                normal_depth = variant_entry['N Cov.']
                normal_fsrb = variant_entry['N FSRB']
                normal_rsrb = variant_entry['N RSRB']
                normal_vaf = variant_entry['N VAF']
                normal_fro = variant_entry['N FRO']
                normal_fao = variant_entry['N FAO']
                transcript_link = chr + "_" + str(pos)
                transcript_id = best_transcript['transcript_id']
                hgvsc = best_transcript['hgvsc']
                canonical = best_transcript['canonical']
                biotype = best_transcript['biotype']
                consequence = ",".join(best_transcript['consequence_terms'])
                ccds = best_transcript['ccds']
                exon = best_transcript['exon']
                protein_id = best_transcript['protein_id']
                hgvsp = best_transcript['hgvsp']
                codons = best_transcript['codons']
                amino_acids = best_transcript['amino_acids']
                strand = best_transcript['strand']
                if not best_transcript['sift_prediction'] == "":
                    sift = "/".join( [ best_transcript['sift_prediction'], str(best_transcript['sift_score']) ] )
                else:
                    sift = ""
                if not best_transcript['polyphen_prediction'] == "":
                    polyphen = "/".join( [ best_transcript['polyphen_prediction'], str(best_transcript['polyphen_score']) ] )
                else:
                    polyphen = ""
                swissprot = best_transcript['swissprot']
                trembl = best_transcript['trembl']
                colocated_link = chr + "_" + str(pos)
                colocated_id = best_colocated_variant['id']
                if 'somatic' in best_colocated_variant.keys():
                    colocated_somatic = best_colocated_variant['somatic']
                else:
                    if 'id' in best_colocated_variant.keys():
                        colocated_somatic = 0
                    else:
                        colocated_somatic = best_colocated_variant['somatic']
                colocated_start = best_colocated_variant['start']
                colocated_end = best_colocated_variant['end']
                colocated_strand = best_colocated_variant['strand']
                colocated_allele = best_colocated_variant['allele_string']
                if 'minor_allele_freq' in best_colocated_variant.keys():
                    colocated_minor_allele_freq = best_colocated_variant['minor_allele_freq']
                elif 'exac_adj_maf' in best_colocated_variant.keys():
                    colocated_minor_allele_freq = best_colocated_variant['exac_adj_maf']
                else:
                    colocated_minor_allele_freq = best_colocated_variant['minor_allele_freq']
    
                if 'minor_allele' in best_colocated_variant.keys():
                    colocated_minor_allele = best_colocated_variant['minor_allele']
                elif 'exac_adj_allele' in best_colocated_variant.keys():
                    colocated_minor_allele = best_colocated_variant['exac_adj_allele']
                else:
                    colocated_minor_allele = best_colocated_variant['minor_allele']
    
                if 'clin_sig' in best_colocated_variant.keys():
                    colocated_clinsig = ",".join(best_colocated_variant['clin_sig'])
                else:
                    colocated_clinsig = best_colocated_variant['clin_sig']
    
                if 'pubmed' in best_colocated_variant.keys():
                    colocated_pubmed = ",".join((str(i) for i in best_colocated_variant['pubmed']))
                else:
                    colocated_pubmed = best_colocated_variant['pubmed']
                
                # convert floats to decimal with 3 decimal points
                
                try:
                    tumor_vaf = "{0:.3f}".format(float(tumor_vaf))
                except ValueError:
                    pass
                try:
                    tumor_fsrb = "{0:.3f}".format(float(tumor_fsrb))
                except ValueError:
                    pass
                try:
                    tumor_rsrb = "{0:.3f}".format(float(tumor_rsrb))
                except ValueError:
                    pass
                try:
                    normal_vaf = "{0:.3f}".format(float(normal_vaf))
                except ValueError:
                    pass
                try:
                    normal_fsrb = "{0:.3f}".format(float(normal_fsrb))
                except ValueError:
                    pass
                try:
                    normal_rsrb = "{0:.3f}".format(float(normal_rsrb))
                except ValueError:
                    pass     
                
                
                values_to_print = [comment,
                                   chr,
                                   gene,
                                   pos,
                                   ref,
                                   alt,
                                   source,
                                   status,
                                   main_consequence,
                                   tumor_depth,
                                   tumor_fsrb,
                                   tumor_rsrb,
                                   tumor_vaf,
                                   tumor_fro,
                                   tumor_fao,
                                   normal_depth,
                                   normal_fsrb,
                                   normal_rsrb,
                                   normal_vaf,
                                   normal_fro,
                                   normal_fao,
                                   transcript_link,
                                   transcript_id,
                                   hgvsc,
                                   canonical,
                                   biotype,
                                   consequence,
                                   ccds,
                                   exon,
                                   protein_id,
                                   hgvsp,
                                   codons,
                                   amino_acids,
                                   strand,
                                   sift,
                                   polyphen,
                                   swissprot,
                                   trembl,
                                   colocated_link,
                                   colocated_id,
                                   colocated_somatic,
                                   colocated_start,
                                   colocated_end,
                                   colocated_strand,
                                   colocated_allele,
                                   colocated_minor_allele_freq,
                                   colocated_minor_allele,
                                   colocated_clinsig,
                                   colocated_pubmed,
                                   assembly_name
                                   ]
    
                return_dict = OrderedDict(zip(header_dict['summary'], values_to_print))
                
                return return_dict
    
            def fix_hyperlinks():
                """Loop through workbook and add hyperlinks to certain fields, if possible."""
                
                summary_sheet = wb.get_sheet_by_name('summary')
                transcript_sheet = wb.get_sheet_by_name('transcripts')
                colocated_sheet = wb.get_sheet_by_name('co-located')
                regulatory_sheet = wb.get_sheet_by_name('regulatory')
    
                # Summary sheet fix
                max_row = summary_sheet.max_row
                for i in range(max_row):
                    row_counter = i+1
                    # skip header
                    if row_counter > 1:
                        # pull coordinates for columns that will be manipulated
                        summary_sheet_chr = summary_sheet.cell(column=2, row=row_counter)
                        summary_sheet_pos = summary_sheet.cell(column=4, row=row_counter)
                        summary_sheet_transcript_link = summary_sheet.cell(column=22, row=row_counter)
                        summary_sheet_colocated_link = summary_sheet.cell(column=39, row=row_counter)
                        
                        # Link from summary sheet to transcript sheet
                        if summary_sheet_transcript_link.value in master_coordinate_dict.keys():                    
                            #transcript_sheet_variant_id.hyperlink = "#summary!%s" % master_coordinate_dict[transcript_sheet_variant_id.value]['summary.transcript link']
                            summary_sheet_transcript_link.value = '=HYPERLINK("#transcripts!%s", "%s")' % (master_coordinate_dict[summary_sheet_transcript_link.value]['transcripts.variant id'], summary_sheet_transcript_link.value)
                            summary_sheet_transcript_link.font = Font(color=BLUE, name="Arial", size=10)
                        
                        # Link from summary sheet to colocated sheet
                        if summary_sheet_colocated_link.value in master_coordinate_dict.keys():                    
                            #transcript_sheet_variant_id.hyperlink = "#summary!%s" % master_coordinate_dict[transcript_sheet_variant_id.value]['summary.transcript link']
                            try:
                                summary_sheet_colocated_link.value = """=HYPERLINK("#'co-located'!%s", "%s")""" % (master_coordinate_dict[summary_sheet_colocated_link.value]['co-located.variant'], summary_sheet_colocated_link.value)
                                summary_sheet_colocated_link.font = Font(color=BLUE, name="Arial", size=10)
                            except KeyError:
                                pass
                        
                        # Add IGV control link
                        tmp_value = summary_sheet_pos.value
                        summary_sheet_pos.value = '=HYPERLINK("http://localhost:60151/goto?locus=%s:%s", "%s")' % (summary_sheet_chr.value, tmp_value, tmp_value)
                        summary_sheet_pos.font = Font(color=BLUE, name="Arial", size=10)
                         
                # Transcript sheet fix
                max_row = transcript_sheet.max_row
                for i in range(max_row):
                    row_counter = i+1
                    transcript_sheet_variant_id = transcript_sheet.cell(column=1, row=row_counter)
                    if transcript_sheet_variant_id.value in master_coordinate_dict.keys():                    
                        #transcript_sheet_variant_id.hyperlink = "#summary!%s" % master_coordinate_dict[transcript_sheet_variant_id.value]['summary.transcript link']
                        transcript_sheet_variant_id.value = '=HYPERLINK("#summary!%s", "%s")' % (master_coordinate_dict[transcript_sheet_variant_id.value]['summary.transcript link'], transcript_sheet_variant_id.value)
                        transcript_sheet_variant_id.font = Font(color=BLUE, name="Arial", size=10)
    
                # Transcript sheet fix
                max_row = colocated_sheet.max_row
                for i in range(max_row):
                    row_counter = i+1
                    colocated_sheet_variant = colocated_sheet.cell(column=1, row=row_counter)
                    if colocated_sheet_variant.value in master_coordinate_dict.keys():                    
                        colocated_sheet_variant.value = '=HYPERLINK("#summary!%s", "%s")' % (master_coordinate_dict[colocated_sheet_variant.value]['summary.co-located link'], colocated_sheet_variant.value)
                        colocated_sheet_variant.font = Font(color=BLUE, name="Arial", size=10)
    
                # Regulatory sheet fix
                max_row = regulatory_sheet.max_row
                for i in range(max_row):
                    row_counter = i+1
                    # skip header
                    if row_counter > 1:
                        # pull coordinates for columns that will be manipulated
                        regulatory_sheet_chr = regulatory_sheet.cell(column=2, row=row_counter)
                        regulatory_sheet_pos = regulatory_sheet.cell(column=4, row=row_counter)
                        regulatory_sheet_colocated_link = regulatory_sheet.cell(column=21, row=row_counter)
    
                        # Link from regulatory sheet to colocated sheet
                        if regulatory_sheet_colocated_link.value in master_coordinate_dict.keys():                    
                            #transcript_sheet_variant_id.hyperlink = "#regulatory!%s" % master_coordinate_dict[transcript_sheet_variant_id.value]['regulatory.transcript link']
                            try:
                                regulatory_sheet_colocated_link.value = """=HYPERLINK("#'co-located'!%s", "%s")""" % (master_coordinate_dict[regulatory_sheet_colocated_link.value]['co-located.variant'], regulatory_sheet_colocated_link.value)
                                regulatory_sheet_colocated_link.font = Font(color=BLUE, name="Arial", size=10)
                            except KeyError:
                                pass
    
                        # Add IGV control link
                        tmp_value = regulatory_sheet_pos.value
                        regulatory_sheet_pos.value = '=HYPERLINK("http://localhost:60151/goto?locus=%s:%s", "%s")' % (regulatory_sheet_chr.value, tmp_value, tmp_value)
                        regulatory_sheet_pos.font = Font(color=BLUE, name="Arial", size=10)
    
    
    
            
            def color_duplicate_entries():
                """Color duplicate entries on the 'summary' sheet."""
                duplicate_variants = defaultdict(int)
                row_counter = 2
                summary_sheet = wb.get_sheet_by_name('summary')
                for row in range(summary_sheet.max_row):
                    row_counter = row+1
                    if row_counter == 1:
                        pass
                    else:
                        chr = summary_sheet.cell(row=row_counter, column=2).value
                        pos = summary_sheet.cell(row=row_counter, column=4).value
                        id = chr + "_" + str(pos)
                        duplicate_variants[id] = duplicate_variants[id] + 1
    
                for row in range(summary_sheet.max_row):
                    row_counter = row+1
                    if row_counter == 1:
                        pass
                    else:
                        chr = summary_sheet.cell(row=row_counter, column=2).value
                        pos = summary_sheet.cell(row=row_counter, column=4).value
                        id = chr + "_" + str(pos)
                        if duplicate_variants[id] > 1:
                            summary_sheet.cell(row=row_counter, column=2).fill = PatternFill(start_color='ff99cc',
                                                                                             end_color='ff99cc',
                                                                                             fill_type='solid'
                                                                                             )
    
            summary_sheet_row_counter, transcript_sheet_row_counter, colocated_sheet_row_counter, regulatory_sheet_row_counter = (2 for i in range(4))
        
            # for storing variant_ids and their coordinates
            # used for fixing links to other pages later
            master_coordinate_dict = defaultdict(dict)
            
            for variant_entry in master_json_list:
                #pp.pprint(variant_entry)
                variant_entry = defaultdict(lambda: "", variant_entry)
                summary_sheet_values_dict = gather_summary_sheet_values(variant_entry)
                summary_sheet_row_counter = write_summary_sheet(summary_sheet_values_dict, summary_sheet_row_counter)
                transcript_sheet_row_counter = write_transcript_sheet(variant_entry, transcript_sheet_row_counter)
                colocated_sheet_row_counter = write_colocated_sheet(variant_entry, colocated_sheet_row_counter)
                regulatory_sheet_row_counter = write_regulatory_sheet(variant_entry, regulatory_sheet_row_counter)
    
            #autofit_columns()
            color_duplicate_entries()
            fix_hyperlinks()
            
        def parse_cnv_files(cnv_tsv, cnv_vcf):
            """Parse IonReporter .tsv and .vcf, and return a dict with all CNV entries."""
            master_cnv_dict = defaultdict(dict)
            with open(cnv_tsv, "r") as f:
                line_counter = 1
                for line in f.readlines():
                    if not line_counter == 1:
                        split_line = line.strip().split("\t")
                        if split_line[2] == "CNV":
                            line_obj = CNV_TSV_Fields(split_line)
                            master_cnv_dict['%s_%s' % (line_obj.chr, line_obj.pos)] = {'comment' : "",
                                                                                       '#chr' : line_obj.chr,
                                                                                       'pos' : line_obj.pos,
                                                                                       'length' : line_obj.length,
                                                                                       'iscn' : line_obj.iscn,
                                                                                       'ploidy' : line_obj.ploidy,
                                                                                       'confidence' : line_obj.confidence,
                                                                                       'precision' : line_obj.precision,
                                                                                       'genes' : line_obj.genes
                                                                                       }
                    line_counter += 1
                f.close()
                
            with open(cnv_vcf, "r") as g:
                line_counter = 1
                for line in g.readlines():
                    if not re.search("^#", line):
                        split_line = line.strip().split("\t")
                        if split_line[4] == "<CNV>":
                            line_obj = CNV_VCF_Fields(split_line, line)
                            master_cnv_dict['%s_%s' % (line_obj.chr, line_obj.pos)].update(
                                                                                           {'numtiles' : line_obj.numtiles,
                                                                                            'end' : line_obj.end
                                                                                            }
                                                                                           )
                g.close()
                
            return master_cnv_dict
        
        def write_cnvs_to_sheet(master_cnv_dict, header_dict):
            """Takes CNV master dict, extracts values, and prints to 'cnv' sheet."""
            
            row_counter = 2
            for cnv_key in natsorted(master_cnv_dict.keys()):
                cnv = defaultdict(lambda: "UNK", master_cnv_dict[cnv_key])
                
                values_to_print = [cnv['comment'],
                                   cnv['#chr'],
                                   cnv['pos'],
                                   cnv['end'],
                                   cnv['length'],
                                   cnv['numtiles'],
                                   cnv['ploidy'],
                                   cnv['iscn'],
                                   cnv['confidence'],
                                   cnv['precision'],
                                   cnv['genes']
                                   ]
                
                print_dict = OrderedDict(zip(header_dict['cnv'], values_to_print))
    
                cnv_sheet = wb.get_sheet_by_name('cnv')
                column_counter = 1
                for k in print_dict.keys():
                    cnv_sheet.cell(row=row_counter, column=column_counter).value = print_dict[k]
                    cnv_sheet.cell(row=row_counter, column=column_counter).font = global_font
                    column_counter += 1
                row_counter += 1
    
        def write_fusions_to_sheet(master_fusion_dict, header_dict):
            """Takes fusion master dict, extracts values, and prints to 'fusion' sheet."""
            
            row_counter = 2
            for fusion_key in OrderedDict(natsorted(master_fusion_dict.iteritems(), key=lambda x: x[1]['Locus'])).keys():
                fusion = defaultdict(lambda: "UNK", master_fusion_dict[fusion_key])
                
                values_to_print = [fusion['Comment'],
                                   fusion['Locus'],
                                   fusion['Type'],
                                   fusion['Genes(Exons)'],
                                   fusion['Read Counts'],
                                   fusion['Oncomine Variant Class'],
                                   fusion['Oncomine Gene Class'],
                                   fusion['Detection'],
                                   fusion["3'/5' Imbalance"],
                                   fusion['COSMIC/NCBI'],
                                   fusion['Variant ID'],
                                   fusion['Read Counts per million']
                                   ]
    
                print_dict = OrderedDict(zip(header_dict['fusion'], values_to_print))
    
                fusion_sheet = wb.get_sheet_by_name('fusion')
                column_counter = 1
                for k in print_dict.keys():
                    fusion_sheet.cell(row=row_counter, column=column_counter).value = print_dict[k]
                    fusion_sheet.cell(row=row_counter, column=column_counter).font = global_font
                    
                    # Special handling for select columns
    #                 if k == "variant id":
    #                     master_coordinate_dict[print_dict['variant id']]['transcripts.variant id'] = transcript_sheet.cell(row=row_counter, column=column_counter).coordinate
    #                 
                    column_counter += 1
                row_counter += 1
    
        def parse_fusion_files(_file):
            """Parses an ionreporter.fusion.vcf file."""
            master_fusion_dict = defaultdict(dict)
            
            svtypes = []
            with open(_file, "r") as f:
                line_counter = 1
                for line in f.readlines():
                    if not re.search("^#", line):
                        line_obj = FusionVCFEntry(line)
                        if line_obj.svtype == "Fusion":
                            if line_obj.fusion.variant_id in master_fusion_dict.keys():
                                master_fusion_dict[line_obj.fusion.variant_id]['Locus'] = master_fusion_dict[line_obj.fusion.variant_id]['Locus'] + " - " + line_obj.fusion.locus
                            else:
                                master_fusion_dict[line_obj.fusion.variant_id] = {'Comment' : "",
                                                                                  'Locus' : line_obj.fusion.locus,
                                                                                  'Type' : line_obj.svtype,
                                                                                  'Genes(Exons)' : line_obj.fusion.genes_and_exons,
                                                                                  'Read Counts' : line_obj.fusion.read_count,
                                                                                  'Oncomine Variant Class' : line_obj.fusion.oncomine_variant_class,
                                                                                  'Oncomine Gene Class' : line_obj.fusion.oncomine_gene_class,
                                                                                  'Detection' : line_obj.fusion.detection,
                                                                                  "3'/5' Imbalance" : "",
                                                                                  'COSMIC/NCBI' : line_obj.fusion.annotation,
                                                                                  'Variant ID' : line_obj.fusion.variant_id,
                                                                                  'Read Counts per million' : line_obj.fusion.rpm
                                                                                  }
                        elif line_obj.svtype == "ExprControl":
                            master_fusion_dict[line_obj.expr_control.variant_id] = {'Comment' : '',
                                                                                    'Locus' : line_obj.expr_control.locus,
                                                                                    'Type' : line_obj.svtype,
                                                                                    'Genes(Exons)' : line_obj.expr_control.genes_and_exons,
                                                                                    'Read Counts' : line_obj.expr_control.read_count,
                                                                                    'Oncomine Variant Class' : line_obj.expr_control.oncomine_variant_class,
                                                                                    'Oncomine Gene Class' : line_obj.expr_control.oncomine_gene_class,
                                                                                    'Detection' : line_obj.expr_control.detection,
                                                                                    "3'/5' Imbalance" : line_obj.expr_control.imbalance_assay,
                                                                                    'COSMIC/NCBI' : line_obj.expr_control.annotation,
                                                                                    'Variant ID' : line_obj.expr_control.variant_id,
                                                                                    'Read Counts per million' : line_obj.expr_control.rpm
                                                                                    }
                        elif line_obj.svtype == "5p3pAssays":
                            master_fusion_dict[line_obj.imbalance_assay.variant_id] = {'Comment' : '',
                                                                                    'Locus' : line_obj.imbalance_assay.locus,
                                                                                    'Type' : line_obj.svtype,
                                                                                    'Genes(Exons)' : line_obj.imbalance_assay.genes_and_exons,
                                                                                    'Read Counts' : line_obj.imbalance_assay.read_count,
                                                                                    'Oncomine Variant Class' : line_obj.imbalance_assay.oncomine_variant_class,
                                                                                    'Oncomine Gene Class' : line_obj.imbalance_assay.oncomine_gene_class,
                                                                                    'Detection' : line_obj.imbalance_assay.detection,
                                                                                    "3'/5' Imbalance" : line_obj.imbalance_assay.imbalance_assay,
                                                                                    'COSMIC/NCBI' : line_obj.imbalance_assay.annotation,
                                                                                    'Variant ID' : line_obj.imbalance_assay.variant_id,
                                                                                    'Read Counts per million' : line_obj.imbalance_assay.rpm
                                                                                    }
                        svtype = re.search("SVTYPE=(\w+?);", line).group(1)
                        svtypes.append(svtype)
                      
            return master_fusion_dict
        
        def parse_specimen_json(_file):
            """Load specimen.json and convert into defaultdict to prepare for printing."""
            try:
                with open(_file, "r") as f:
                    file_loaded = json.loads(f.read())
                    master_specimen_dict = defaultdict(lambda: "Unknown", file_loaded)
            except:
                master_specimen_dict = defaultdict(lambda: "Unknown")
    
            return master_specimen_dict
        
        def write_cnv_sheet_headers():
            
    
            cnv_sheet_headers = ['comment',
                                 '#chr',
                                 'pos',
                                 'end',
                                 'length',
                                 'numtiles',
                                 'ploidy',
                                 'iscn',
                                 'confidence',
                                 'precision',
                                 'genes'
                                 ]
            
            cnv_sheet = wb.get_sheet_by_name('cnv')
            row_counter = 1
            for header in cnv_sheet_headers:
                cnv_sheet.cell(row=row_counter, column=cnv_sheet_headers.index(header)+1).value = header

            tmp_dict = {'cnv' : cnv_sheet_headers}

            return tmp_dict
        
        def write_fusion_sheet_headers():
            
            fusion_sheet_headers = ['Comment',
                                   'Locus',
                                   'Type',
                                   'Genes(Exons)',
                                   'Read Counts',
                                   'Oncomine Variant Class',
                                   'Oncomine Gene Class',
                                   'Detection',
                                   "3'/5' Imbalance",
                                   'COSMIC/NCBI',
                                   'Variant ID',
                                   'Read Counts per million']
            
            fusion_sheet = wb.get_sheet_by_name('fusion')
            row_counter = 1
            for header in fusion_sheet_headers:
                fusion_sheet.cell(row=row_counter, column=fusion_sheet_headers.index(header)+1).value = header

            tmp_dict = {'fusion' : fusion_sheet_headers}

            return tmp_dict
        
        def write_specimen_sheet(master_specimen_dict):
            active_sheet = wb.get_sheet_by_name("Specimen")
            
            print_dict = OrderedDict([
                                      ("Pipeline Version", master_specimen_dict['pipeline_version']),
                                      ("Patient", master_specimen_dict['patient']),
                                      ("Copath#", master_specimen_dict['Copath#']),
                                      ("Dissection", master_specimen_dict['dissection']),
                                      ("Malignant cells", master_specimen_dict['malignant_cells']),
                                      ("Tumor", master_specimen_dict['tumor']),
                                      ("Normal", master_specimen_dict['normal']),
                                      ("Panel", master_specimen_dict['panel']),
                                      ("Requested by", master_specimen_dict['requested_by']),
                                      ("Sequencing done", master_specimen_dict['sequencing_done']),
                                      ("Insufficient RNA for Fusions", master_specimen_dict['insufficient_RNA'])
                                      ])
            
            row_counter = 1
            for k in print_dict.keys():
                active_sheet.cell(column=1, row=row_counter).value = k
                active_sheet.cell(column=1, row=row_counter).font = global_font
                active_sheet.cell(column=2, row=row_counter).value = print_dict[k]
                active_sheet.cell(column=2, row=row_counter).font = global_font
                row_counter += 1
            
        
        # FILE DETECTION AND SOME GLOBAL VARIABLE DECLARATIONS
        
        if args.autodetect is True: 
            detected_files = validate_input()
        else:
            detected_files = []
            command_line_file_options = [args.variant_jsons,
                                         args.ir_cnv_tsv,
                                         args.ir_cnv_vcf,
                                         args.specimen_json,
                                         args.ir_fusion_vcf]
            
            command_line_file_options = (x for x in list(traverse(command_line_file_options)) if x != None)
            for _file in command_line_file_options:
                detected_files.append(_file)
        
        pp = pprint.PrettyPrinter(indent=4)
    
        wb = open_output_xlsx()
        global_font = Font(name="Arial", size="10")
        create_variant_sheets()
         
        
    
        print "-----------------------------------"
        print "---------VARIANT PROCESSING--------"
        print "-----------------------------------"
    
        master_json_list = []
        detected_jsons = []
        for f in detected_files:
            if f.endswith(".json") and not f.endswith(".specimen.json"):
                detected_jsons.append(f)
                print f
                parse_variant_json(f)
        if detected_jsons:
            print "OK: DETECTED %s variant .json files" % len(detected_jsons)
            print "OK: Processing the following .json files..."
            for k in detected_jsons:
                print "OK: (%s) %s" % (detected_jsons.index(k)+1, k)
            header_dict = write_variant_sheets_headers() 
            write_variants()
        else:
            print "FAIL: Did not detect any variant .json files"
        
        # CNV PARSING AND WRITING
        cnv_files = {}
        for f in detected_files:
            if f.endswith(".ionreporter.tsv"):
                cnv_files['IonReporter CNV .tsv'] = f
            if f.endswith(".ionreporter.cnv.vcf"):
                cnv_files['IonReporter CNV .vcf'] = f
    
        print "-----------------------------------"
        print "-----------CNV PROCESSING----------"
        print "-----------------------------------"
        
        if not cnv_files:
            print "WARNING: No CNV files detected.  Assuming there are no CNVs for this batch of files..."
        else:
            necessary_keys = ['IonReporter CNV .tsv', 'IonReporter CNV .vcf']
            for k in necessary_keys:
                if k not in cnv_files.keys():
                    sys.exit("ERROR: Both IonReporter CNV .tsv and .vcf files are needed for CNV parsing.  Exiting...")
                else:
                    print "OK: %s detected" % k
            
            create_sheet('cnv')
            header_dict = write_cnv_sheet_headers()
            master_cnv_dict = parse_cnv_files(cnv_files['IonReporter CNV .tsv'], cnv_files['IonReporter CNV .vcf'])
            write_cnvs_to_sheet(master_cnv_dict, header_dict)
    
    
    
        # FUSION PARSING AND WRITING
    
    
        print "-----------------------------------"
        print "---------FUSION PROCESSING---------"
        print "-----------------------------------"
        
        detected_fusion_bool = []
        for f in detected_files:
            if f.endswith(".ionreporter.fusions.vcf"):
                detected_fusion_bool.append(True)
                fusion_vcf = f
                create_sheet('fusion')
                header_dict = write_fusion_sheet_headers()
                master_fusion_dict = parse_fusion_files(fusion_vcf)
                write_fusions_to_sheet(master_fusion_dict, header_dict)
            else:
                detected_fusion_bool.append(False)
                
        if True in detected_fusion_bool:
            print "OK: Detected fusions.vcf"
        else:
            print "WARNING: No fusions.vcf detected.  Fusions will not be parsed.  Continuing..."
                
                
                
        # SPECIMEN INFO PARSING
    
    
        print "-----------------------------------"
        print "--------SPECIMEN PROCESSING--------"
        print "-----------------------------------"
        
        detected_specimen_bool = []
        for f in detected_files:
            if f.endswith(".specimen.json"):
                detected_specimen_bool.append(True)
                specimen_json = f
            else:
                detected_specimen_bool.append(False)
        
        if True in detected_specimen_bool:
            print "OK: Detected specimen.json"
            master_specimen_dict = parse_specimen_json(specimen_json)
            write_specimen_sheet(master_specimen_dict)
        else:
            master_specimen_dict = parse_specimen_json(None)
            write_specimen_sheet(master_specimen_dict)
            print "WARNING: No specimen.json file detected.  Specimen info will not be parsed.  Continuing..."
    
    
        # Autofit columns
        autofit_columns()
        # Save workbook
        wb.save(filename="%s.xlsx" % args.output_basename[0])      
        
    
    
    tn_pipeline_output_processing()
    
if __name__ == "__main__":
    main()