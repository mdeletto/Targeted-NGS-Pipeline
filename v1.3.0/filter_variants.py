#!/usr/bin/python

import openpyxl, re, optparse, sys
from openpyxl.styles import Font, Fill
from openpyxl.styles.colors import BLUE
from openpyxl.utils import get_column_letter
import pprint
from collections import defaultdict
#-------------------------------------------------------------------------------------------
#----------------------------Command line parser and arguments------------------------------
#-------------------------------------------------------------------------------------------

desc="""This script automatically applies filters to the DownstreamReporting spreadsheet.  At the time of authorship, filters are only set up for targeting sequencing panels only."""
vers="1.0"

parser = optparse.OptionParser(description=desc, version=vers)

parser.add_option('-i', help='DownstreamReporting spreadsheet input (.xlsx only)', dest='input', action='store')
parser.add_option('-p', help='Panel type <OCP,CCP,BRCA>', dest='panel', action='store', default=None) # This option is not currently used in the program, but will in the future to apply panel-specific filters.
parser.add_option('-o', help='DownstreamReporting filtered output for reupload', dest='output', action='store')
parser.add_option('--minimum-vaf', help='Minimum variant allele frequency for detected variants', dest='minimum_vaf', action='store', default=0.10)

monthly_qc_args = optparse.OptionGroup(parser, "MONTHLY QC",
                             "Options only to be used when analyzing monthly Acrometrix MultiMixD Hotspot Control.  Enabling these options adds significant runtime.  This needs to be optimized")

monthly_qc_args.add_option("--enable_monthly_qc", action="store_true", help="Flag to enable monthly QC check", dest="enable_monthly_qc", default=False)
monthly_qc_args.add_option("--qc_bed", action="store", help="Path to custom Acrometrix Hotspot Control bed file.  Default is custom OCP143 intersection w/ Acrometrix Hotspot Control", dest="qc_bed", default="/home/michael/Cases/QC/reference_files/OCP143.acrometrix_oncology_hotspot_control.multimixD.synonymous_included.recommended.bed")

parser.add_option_group(monthly_qc_args)

(opts, args) = parser.parse_args()

def parse_command_line(opts):
    mandatory_options = ['input','output']
    for m in mandatory_options:
        # Making sure all mandatory options appeared
        if not opts.__dict__[m]:
            print "Mandatory option is missing!\n"
            parser.print_help()
            sys.exit()
    
    if not opts.__dict__['input'].endswith(".xlsx") or not opts.__dict__['output'].endswith(".xlsx"):
        print "Formats other than .xlsx are not supported!"
        parser.print_help()
        sys.exit()
        
    if opts.enable_monthly_qc is True:
        message = "PERFORMING MONTHLY QC CHECK.  PLEASE WAIT....."
        print "!"+"-"*len(message)+"!"
        print message

class Sheet_name_and_column_slice_matrix:
    def __init__(self,sheet):
        self.sheet_name = sheet[0]
        self.header_slice_start = sheet[1]
        self.header_slice_end = sheet[2]

class Summary_headers_indices:
    def __init__(self,headers_index_dict):
        self.comment = headers_index_dict["comment"]
        self.chr = headers_index_dict["#chr"]
        self.gene = headers_index_dict["gene"]
        self.pos = headers_index_dict["pos"]
        self.ref = headers_index_dict["ref"]
        self.alt = headers_index_dict["alt"]
        self.source = headers_index_dict["source"]
        self.status = headers_index_dict["status"]
        self.main_consequence = headers_index_dict["main consequence"]
        self.tumor_coverage = headers_index_dict["T Cov."]
        self.tumor_forward_strand_read_bias = headers_index_dict["T FSRB"]
        self.tumor_reverse_strand_read_bias = headers_index_dict["T RSRB"]
        self.tumor_VAF = headers_index_dict["T VAF"]
        self.tumor_RO = headers_index_dict["T FRO"]
        self.tumor_AO = headers_index_dict["T FAO"]
        self.normal_coverage = headers_index_dict["N Cov."]
        self.normal_forward_strand_read_bias = headers_index_dict["N FSRB"]
        self.normal_reverse_strand_read_bias = headers_index_dict["N RSRB"]
        self.normal_VAF = headers_index_dict["N VAF"]
        self.normal_RO = headers_index_dict["N FRO"]
        self.normal_AO = headers_index_dict["N FAO"] 
        self.transcript_link = headers_index_dict["transcript link"]
        self.transcript_id = headers_index_dict["transcript id"]
        self.hgvsc = headers_index_dict["hgvsc"]
        self.canonical = headers_index_dict["canonical"]
        self.biotype = headers_index_dict["biotype"]
        self.consequence = headers_index_dict["consequence"]
        self.ccds = headers_index_dict["ccds"]
        self.exon = headers_index_dict["exon"]
        self.protein_id = headers_index_dict["protein id"]
        self.hgvsp = headers_index_dict["hgvsp"]
        self.codons = headers_index_dict["codons"]
        self.amino_acids = headers_index_dict["amino acids"]
        self.strand = headers_index_dict["strand"]
        self.sift = headers_index_dict["sift"]
        self.polyphen = headers_index_dict["polyphen"]
        self.swissprot = headers_index_dict["swissprot"]
        self.trembl = headers_index_dict["trembl"]
        self.colocated_link = headers_index_dict["co-located link"]
        self.colocated_id = headers_index_dict["co-located id"]
        self.colocated_somatic = headers_index_dict["co-located somatic"]
        self.colocated_start = headers_index_dict["cl-start"]
        self.colocated_end = headers_index_dict["cl-end"]
        self.colocated_strand = headers_index_dict["cl-strand"]
        self.colocated_allele = headers_index_dict["cl-allele"]
        self.colocated_minor_allele_freq = headers_index_dict["cl-minor allele freq"]
        self.colocated_minor_allele = headers_index_dict["cl-minor allele"]
        self.colocated_clin_sig = headers_index_dict["cl-clin sig"]
        self.colocated_pubmed = headers_index_dict["cl-pubmed"]

class Transcript_headers_indices:
    def __init__(self,headers_index_dict):
        self.variant_id = headers_index_dict["variant id"]
        self.gene = headers_index_dict["gene"]
        self.transcript_id = headers_index_dict["transcript id"]
        self.hgvsc = headers_index_dict["hgvsc"]
        self.canonical = headers_index_dict["canonical"]
        self.biotype = headers_index_dict["biotype"]
        self.consequence = headers_index_dict["consequence"]
        self.ccds = headers_index_dict["ccds"]
        self.exon = headers_index_dict["exon"]
        self.protein_id = headers_index_dict["protein id"]
        self.hgvsp = headers_index_dict["hgvsp"]
        self.codons = headers_index_dict["codons"]
        self.amino_acids = headers_index_dict["amino acids"]
        self.strand = headers_index_dict["strand"]
        self.sift = headers_index_dict["sift"]
        self.polyphen = headers_index_dict["polyphen"]
        self.swissprot = headers_index_dict["swissprot"]
        self.trembl = headers_index_dict["trembl"]

class Colocated_headers_indices:
    def __init__(self,headers_index_dict):
        self.variant = headers_index_dict["variant"]
        self.known_id = headers_index_dict["known id"]
        self.somatic = headers_index_dict["somatic"]
        self.start = headers_index_dict["start"]
        self.end = headers_index_dict["end"]
        self.strand = headers_index_dict["strand"]
        self.allele = headers_index_dict["allele"]
        self.minor_allele_freq = headers_index_dict["minor allele freq"]
        self.minor_allele = headers_index_dict["minor allele"]
        self.clin_sig = headers_index_dict["clin sig"]
        self.pubmed = headers_index_dict["pubmed"]

class CNV_headers_indices:
    def __init__(self,headers_index_dict):
        self.comment = headers_index_dict["comment"]
        self.chr = headers_index_dict["#chr"]
        self.pos = headers_index_dict["pos"]
        self.end = headers_index_dict["end"]
        self.length = headers_index_dict["length"]
        self.numtiles = headers_index_dict["numtiles"]
        self.ploidy = headers_index_dict["ploidy"]
        self.iscn = headers_index_dict["iscn"]
        self.confidence = headers_index_dict["confidence"]
        self.precision = headers_index_dict["precision"]
        self.genes = headers_index_dict["genes"]
        
class Fusion_headers_indices:
    def __init__(self,headers_index_dict):
        self.comment = headers_index_dict["Comment"]
        self.locus = headers_index_dict["Locus"]
        self.type = headers_index_dict["Type"]
        self.gene_exon_boundaries = headers_index_dict["Genes(Exons)"]
        self.read_counts = headers_index_dict["Read Counts"]
        self.oncomine_variant_class = headers_index_dict["Oncomine Variant Class"]
        self.oncomine_gene_class = headers_index_dict["Oncomine Gene Class"]
        self.detection = headers_index_dict["Detection"]
        self.imbalance_value = headers_index_dict["3'/5' Imbalance"]
        self.annotation = headers_index_dict["COSMIC/NCBI"]
        self.variant_id = headers_index_dict["Variant ID"]

def define_header_indices(wb,sheet_name,header_slice_start_column,header_slice_end_column):
    try:
        sheet = wb.get_sheet_by_name(sheet_name)
        headers = sheet[header_slice_start_column:header_slice_end_column]
        headers_index_dict = {}
        for column in headers:
            column_counter = 0
            for cell in column:
                column_counter += 1
                headers_index_dict[cell.value]=column_counter
        return headers_index_dict
    except:
        print "'%s' sheet was not processed (it probably doesn't exist in the spreadsheet!)" % sheet_name
        return 0

def recreate_transcript_hyperlinks(wb,sheetname,indices,transcript_id_row_dict):
    
    transcript_sheet = wb.get_sheet_by_name(sheetname)
    row_counter = 0
    
    for row in transcript_sheet.rows:
        row_counter+=1
        for cell in row:
            cell.font = Font(name="Liberation Sans",size=10) # Apply font to all cells on sheet
        if row_counter==1: # pass for header
            pass
        else:
            variant_id = transcript_sheet.cell(row=row_counter,column=indices.variant_id)
            gene = transcript_sheet.cell(row=row_counter,column=indices.gene)
#             if variant_id.value in transcript_id_row_dict.keys():
#                 pass
#             else:
            transcript_id_row_dict[variant_id.value]=variant_id.coordinate
            gene.hyperlink = "http://www.genecards.org/cgi-bin/carddisp.pl?gene=%s" % (gene.value)
            
        
    return transcript_id_row_dict    

def recreate_colocated_hyperlinks(wb,sheetname,indices,colocated_id_row_dict):
    
    colocated_sheet = wb.get_sheet_by_name(sheetname)
    row_counter = 0
    
    for row in colocated_sheet.rows:
        row_counter+=1
        for cell in row:
            cell.font = Font(name="Liberation Sans",size=10) # Apply font to all cells on sheet
        if row_counter==1: # pass for header
            pass
        else:
            variant = colocated_sheet.cell(row=row_counter,column=indices.variant)
#             if variant.value in colocated_id_row_dict.keys():
#                 pass
#             else:
            colocated_id_row_dict[variant.value]=variant.coordinate
    
    return colocated_id_row_dict    

def apply_summary_filters(wb,sheetname,indices,transcript_id_row_dict,colocated_id_row_dict):

    summary_sheet = wb.get_sheet_by_name(sheetname)
    transcript_sheet = wb.get_sheet_by_name("transcripts")
    colocated_sheet = wb.get_sheet_by_name("co-located")

    variant_summary_dict = {}

    row_counter = 0
    for row in summary_sheet.rows:
        row_counter += 1
        for cell in row:
            cell.font = Font(name="Liberation Sans",size=10)        
            #cell.font = Font(color="001B49")    
        if row_counter==1: # pass for header
            pass
        else:
            
            status = summary_sheet.cell(row=row_counter,column=indices.status)
            comment = summary_sheet.cell(row=row_counter,column=indices.comment)
            main_consequence = summary_sheet.cell(row=row_counter,column=indices.main_consequence)
            alt = summary_sheet.cell(row=row_counter,column=indices.alt)
            tumor_VAF = summary_sheet.cell(row=row_counter,column=indices.tumor_VAF)
            normal_VAF = summary_sheet.cell(row=row_counter,column=indices.normal_VAF)
            transcript_link = summary_sheet.cell(row=row_counter,column=indices.transcript_link)
            colocated_link = summary_sheet.cell(row=row_counter,column=indices.colocated_link)
            colocated_variant = summary_sheet.cell(row=row_counter,column=indices.colocated_id)
            colocated_minor_allele_freq = summary_sheet.cell(row=row_counter,column=indices.colocated_minor_allele_freq)
            colocated_clin_sig = summary_sheet.cell(row=row_counter,column=indices.colocated_clin_sig)
            chrom = summary_sheet.cell(row=row_counter,column=indices.chr)
            pos = summary_sheet.cell(row=row_counter,column=indices.pos)
            
            hgvsc = summary_sheet.cell(row=row_counter,column=indices.hgvsc)
            hgvsp = summary_sheet.cell(row=row_counter,column=indices.hgvsp)
            gene = summary_sheet.cell(row=row_counter,column=indices.gene)
            coordinate = comment.coordinate
            
            if status.value == "Somatic":
                if (main_consequence.value == "3_prime_UTR_variant") or (main_consequence.value == "5_prime_UTR_variant") \
                or (main_consequence.value == "intron_variant") or (main_consequence.value == "non_coding_transcript_exon_variant") or (main_consequence.value == "synonymous_variant"):
                    summary_sheet[coordinate]="non-deleterious change(not reviewed)"
                elif tumor_VAF.value is None:
                    pass
                elif tumor_VAF.value == "UNK":
                    pass
                elif float(tumor_VAF.value) <= float(opts.minimum_vaf):
                    summary_sheet[coordinate]="no significant VAF in tumor"
                else:
                    pass
            elif status.value == "Germline":
                if (main_consequence.value == "3_prime_UTR_variant") or (main_consequence.value == "5_prime_UTR_variant") \
                or (main_consequence.value == "intron_variant") or (main_consequence.value == "non_coding_transcript_exon_variant") or (main_consequence.value == "synonymous_variant"):
                    summary_sheet[coordinate]="non-deleterious change(not reviewed)"
                elif tumor_VAF.value is None:
                    pass
                elif float(tumor_VAF.value) <= 0.35:
                    summary_sheet[coordinate]="no significant VAF in tumor"               
                elif re.search("^rs",str(colocated_variant.value)):
                    try:
                        if colocated_minor_allele_freq.value is None:
                            pass
                        elif float(colocated_minor_allele_freq.value) >= 0.001:
                            if re.search("pathogenic",str(colocated_clin_sig.value)):
                                pass
                            else:
                                summary_sheet[coordinate]="known SNP"
                        else:
                            pass
                    except Exception,e:
                        print str(e)
                else:
                    pass   
            elif status.value == "LOH":
                if (main_consequence.value == "3_prime_UTR_variant") or (main_consequence.value == "5_prime_UTR_variant") \
                or (main_consequence.value == "intron_variant") or (main_consequence.value == "non_coding_transcript_exon_variant") or (main_consequence.value == "synonymous_variant"):
                    summary_sheet[coordinate]="non-deleterious change(not reviewed)"
                elif tumor_VAF.value is None:
                    pass
                elif float(tumor_VAF.value) <= float(opts.minimum_vaf):
                    summary_sheet[coordinate]="no significant VAF in tumor"
                elif float(tumor_VAF.value) <= 0.50 and float(normal_VAF.value) <= 0.50:
                    summary_sheet[coordinate]="also in normal"
                else:
                    pass

            # Add IGV hyperlink
            pos_as_int = int(pos.value)
            pos.value = '=HYPERLINK("http://localhost:60151/goto?locus=%s:%s", "%s")' % (chrom.value, str(int(pos.value)), str(int(pos.value)))
            pos.font = Font(name="Liberation Sans",size=10,color=BLUE)

            for key in transcript_id_row_dict.keys():
                if transcript_link.value == key:
                    transcript_link.value = '=HYPERLINK("#transcripts!%s", "%s")' % (transcript_id_row_dict[key], key)
                    transcript_link.font = Font(name="Liberation Sans",size=10,color=BLUE)
                    transcript_sheet['%s' % transcript_id_row_dict[key]].value = '=HYPERLINK("#summary!%s", "%s")' % (transcript_link.coordinate, key)
                    transcript_sheet['%s' % transcript_id_row_dict[key]].font = Font(name="Liberation Sans",size=10,color=BLUE)
                else:
                    pass
                   
            for key in colocated_id_row_dict.keys():
                if colocated_link.value == key:
                    colocated_link.value = '=HYPERLINK("#\'co-located\'!%s", "%s")' % (colocated_id_row_dict[key], key)
                    colocated_link.font = Font(name="Liberation Sans",size=10,color=BLUE)
                    colocated_sheet['%s' % colocated_id_row_dict[key]].value = '=HYPERLINK("#summary!%s", "%s")' % (colocated_link.coordinate, key)
                    colocated_sheet['%s' % colocated_id_row_dict[key]].font = Font(name="Liberation Sans",size=10,color=BLUE)
                else:
                    pass


            # Extract fields for variant summary dict.  We can use it parse out values later
            extracted_fields = {'status' : status.value,
                                'comment' : comment.value,
                                'comment_coordinate' : comment.coordinate,
                                'main consequence' : main_consequence.value,
                                'gene' : gene.value,
                                'T VAF' : tumor_VAF.value,
                                'N VAF' : normal_VAF.value, 
                                'transcript link' : transcript_link.value,
                                'co-located link' : colocated_link.value, 
                                'co-located id' : colocated_variant.value, 
                                'cl-minor allele freq' : colocated_minor_allele_freq.value, 
                                'cl-clin sig' : colocated_clin_sig.value,
                                '#chr' : chrom.value,
                                'pos' : pos_as_int, #pos.value,
                                'hgvsp' : hgvsp.value,
                                'hgvsc' : hgvsc.value
                                }

            #print '%s_%s_%s' % (chrom.value, pos_as_int, alt.value)
            variant_summary_dict['%s_%s_%s' % (chrom.value, pos_as_int, alt.value)] = extracted_fields

    return variant_summary_dict

def apply_CNV_filter(wb,sheetname,indices):

    cnv_sheet = wb.get_sheet_by_name(sheetname)
    row_counter = 0
    confidence_sort_dict = {} 
    for row in cnv_sheet.rows:
        row_counter+=1
        for cell in row:
            cell.font = Font(name="Liberation Sans",size=10)         
        if row_counter==1: # pass for header
            pass 
        else:
            confidence_score = cnv_sheet.cell(row=row_counter,column=indices.confidence)
            confidence_sort_dict[row] = float(confidence_score.value)
    
    
    new_coordinate_key_value = {}
    row_counter = 1
    for keys in sorted(confidence_sort_dict, key=confidence_sort_dict.get, reverse=True):
        row_counter += 1
        column_counter = 0
        for cell in keys:
            column_counter += 1
            value = cell.value
            column_letter = colnum_string(column_counter)
            coordinate = str(column_letter)+str(row_counter)
            new_coordinate_key_value[coordinate] = value
    
    for key in new_coordinate_key_value.keys():
        cnv_sheet[key] = new_coordinate_key_value[key]

    row_counter = 0
    for row in cnv_sheet.rows:
        row_counter+=1
        if row_counter == 1:
            pass
        else:
            confidence_score = cnv_sheet.cell(row=row_counter,column=indices.confidence)
            precision_score = cnv_sheet.cell(row=row_counter,column=indices.precision)
            comment = cnv_sheet.cell(row=row_counter,column=indices.comment)
            if float(confidence_score.value) < 20:
                comment.value = "Confidence score less than 20"
                for cell in row:
                    cell.font = Font(name="Liberation Sans",size=10,color='909090')
            else:
                if float(precision_score.value) < 20:
                    comment.value = "OK: Low precision"
                    precision_score.font = Font(name="Liberation Sans",size=10,color='909090')
                else:
                    comment.value = "OK"

def apply_fusion_filter(wb,sheetname,indices):
    fusion_sheet = wb.get_sheet_by_name(sheetname)
    row_counter = 0
    fusion_read_count_sort_dict = {} # only 'Fusion' calls
    prime_assay_rows = []
    control_rows = []
    gene_expression_rows = []
    
    for row in fusion_sheet.rows:
        row_counter+=1
        for cell in row:
            cell.font = Font(name="Liberation Sans",size=10)         
        if row_counter==1: # pass for header
            pass 
        else:
            fusion_type = fusion_sheet.cell(row=row_counter,column=indices.type)
            read_counts = fusion_sheet.cell(row=row_counter,column=indices.read_counts)
            if fusion_type.value == "Fusion":
                fusion_read_count_sort_dict[row]=read_counts.value
            elif fusion_type.value == "5p3pAssays":
                prime_assay_rows.append(row)
            elif fusion_type.value == "ExprControl":
                control_rows.append(row)
            elif fusion_type.value == "GeneExpression":
                gene_expression_rows.append(row)
            else:
                print "Unsupported fusion type"

    new_coordinate_key_value = {}
    row_counter = 1
    for keys in sorted(fusion_read_count_sort_dict, key=fusion_read_count_sort_dict.get, reverse=True):
        row_counter += 1
        column_counter = 0
        for cell in keys:
            column_counter += 1
            value = cell.value
            column_letter = colnum_string(column_counter)
            coordinate = str(column_letter)+str(row_counter)
            new_coordinate_key_value[coordinate] = value
    for row in prime_assay_rows:
        row_counter += 1
        column_counter = 0
        for cell in row:
            column_counter += 1
            value = cell.value
            column_letter = colnum_string(column_counter)
            coordinate = str(column_letter)+str(row_counter)
            new_coordinate_key_value[coordinate] = value
    for row in control_rows:
        row_counter += 1
        column_counter = 0
        for cell in row:
            column_counter += 1
            value = cell.value
            column_letter = colnum_string(column_counter)
            coordinate = str(column_letter)+str(row_counter)
            new_coordinate_key_value[coordinate] = value
    for row in gene_expression_rows:
        row_counter += 1
        column_counter = 0
        for cell in row:
            column_counter += 1
            value = cell.value
            column_letter = colnum_string(column_counter)
            coordinate = str(column_letter)+str(row_counter)
            new_coordinate_key_value[coordinate] = value
    
    for key in new_coordinate_key_value.keys():
        fusion_sheet[key] = new_coordinate_key_value[key]
        
    row_counter = 0
    for row in fusion_sheet.rows:
        row_counter+=1
        if row_counter == 1:
            pass
        else:
            fusion_type = fusion_sheet.cell(row=row_counter,column=indices.type)
            read_counts = fusion_sheet.cell(row=row_counter,column=indices.read_counts)
            comment = fusion_sheet.cell(row=row_counter,column=indices.comment)
            
            if fusion_type.value == "Fusion":
                if int(read_counts.value) >= 20:
                    comment.value = "OK: Fusion"
                else:
                    for cell in row:
                        cell.font = Font(name="Liberation Sans",size=10,color='909090')
                    comment.value = "Absent"
            elif fusion_type.value == "ExprControl" or fusion_type.value == "GeneExpression":
                for cell in row:
                    cell.font = Font(name="Liberation Sans",size=10,color='909090')
                if int(read_counts.value) >= 20:
                    comment.value = "Control: Present"
                else:
                    comment.value = "Control: Absent"
            elif fusion_type.value == "5p3pAssays":
                for cell in row:
                    cell.font = Font(name="Liberation Sans",size=10,color='909090')
                comment.value = "5p3pAssay"
 

def colnum_string(n):
    div=n
    string=""
    temp=0
    while div>0:
        module=(div-1)%26
        string=chr(65+module)+string
        div=int((div-module)/26)
    return string   
  
def perform_monthly_QC_check(wb, variant_summary_dict):
    
    def read_and_parse_QC_bed():
    
        class QC_bed_details:
            def __init__(self,split_line):
                
                def parse_details(details):
                    details_dict = defaultdict(str)
                    split_details = details.strip().split(";")
                    for split_detail in filter(None,split_details):
                        split_detail = split_detail.split("=", 1)
                        details_dict[split_detail[0]] = split_detail[1]
                    
                    return details_dict
                
                self.chrom = split_line[0]
                self.start = split_line[1]
                self.end = split_line[2]
                self.details = split_line[3]
                #self.strand = split_line[5]
    
        
                details_dict = parse_details(self.details)
                
                self.mutation_cds = details_dict['MUTATION_CDS']
                self.mutation_aa = details_dict['MUTATION_AA']
                self.ref_allele = details_dict['REF']
                self.target_frequency = details_dict['TARGET_FREQUENCY']
                self.alt_allele = details_dict['ALT']
                #self.length = details_dict['LENGTH']
                self.gene = details_dict['GENE']
                self.mutation_id = details_dict['MUTATION_ID']
                self.mutation_type = details_dict['TYPE']
                self.amplicon_id = details_dict['AMPLICON']
                self.details_dict = details_dict
        
        QC_coordinate_detail_dict = defaultdict(str)
        #for line in open(QC_bed,"r").readlines():
        for line in open(opts.qc_bed,"r").readlines():
            split_line = line.strip().split()
            qc_bed_details = QC_bed_details(split_line)
            # WARNING: The following conditional is a quick fix to only use CHP* amplicons for the CHPv2/HSM monthly QC
            if not re.search("CHP", qc_bed_details.amplicon_id) and (opts.panel=="CHPv2" or opts.panel == "HSM"):
                pass
            else:
                #QC_coordinate_detail_dict['%s_%s_%s' % (qc_bed_details.chrom, qc_bed_details.start, qc_bed_details.alt_allele)] = qc_bed_details.details_dict
                QC_coordinate_detail_dict['%s_%s_%s' % (qc_bed_details.chrom, qc_bed_details.end, qc_bed_details.alt_allele)] = qc_bed_details.details_dict
            
        return QC_coordinate_detail_dict
    
    def create_qc_sheet(wb):
        qc_sheet = wb.create_sheet()
        qc_sheet.title = "qc"
        qc_sheet = wb.get_sheet_by_name("qc")

        return qc_sheet
    
    def write_qc_header(qc_sheet):
        
        headers = ["comment",
           "pass/fail",
           "#chr",
           "gene",
           "pos",
           "ref",
           "alt",
           "mutation type",
           "status",
           "main consequence",
           "T VAF",
           "N VAF",
           "expected VAF",
           "transcript link",
           "expected HGVSc",
           "expected HGVSp",
           "mutation ID"
           ]
        
        row_counter = 1
        column_counter = 1 
        for header in headers:
            cell = qc_sheet.cell(row=row_counter, column=column_counter)
            cell.value = header
            
            if qc_sheet.column_dimensions[get_column_letter(column_counter)].width < len(str(cell.value)):
                qc_sheet.column_dimensions[get_column_letter(column_counter)].width = len(str(cell.value)) + 5
            
            column_counter += 1
        
        row_counter += 1
        
        return row_counter
    
    def write_qc_variants(variant_summary_dict, qc_dict, row_counter, qc_sheet, summary_sheet):
        
        
        def expected_observed_VAF_test(qc_dict_key,variant_dict_key):
            
            if qc_dict[qc_dict_key]['TARGET_FREQUENCY'] == 'genomic' or qc_dict[qc_dict_key]['TARGET_FREQUENCY'] == '':
                comment = "DETECTED"
                pass_or_fail = "PASS"
            else:
                frequency_range = qc_dict[qc_dict_key]['TARGET_FREQUENCY'].strip("%").split("-")
                lower_frequency_bound, upper_frequency_bound = (float(float(i)/100) for i in frequency_range)
                if float(variant_summary_dict[variant_dict_key]['T VAF']) > upper_frequency_bound or float(variant_summary_dict[variant_dict_key]['T VAF']) < lower_frequency_bound:
                    comment = "WARNING: Detected but not within expected VAF range"
                    pass_or_fail = "PASS-W"
                else:
                    comment = "DETECTED"
                    pass_or_fail = "PASS"
            
            return_list = []
            return_list.append(comment)
            return_list.append(pass_or_fail)
            
            return return_list
        
        def select_final_qc_fields_from_variant_summary_dict(k):
            try:
                chrom = variant_summary_dict[k]['#chr']
                pos = variant_summary_dict[k]['pos']
                status = variant_summary_dict[k]['status']
                main_consequence = variant_summary_dict[k]['main consequence']
                t_vaf = variant_summary_dict[k]['T VAF']
                n_vaf = variant_summary_dict[k]['N VAF']
                transcript_link = variant_summary_dict[k]['transcript link']
            except KeyError:
                chrom, pos, alt_allele = k.split("_")
                status, main_consequence, t_vaf, n_vaf, transcript_link = ("" for i in range(5))
            
            selected_fields = [chrom, pos, status, main_consequence, 
                               t_vaf, n_vaf, transcript_link]
            
            return selected_fields
            
        def select_final_qc_fields_from_qc_dict(k):
            
            gene = qc_dict[k]['GENE']
            ref = qc_dict[k]['REF']
            alt = qc_dict[k]['ALT']
            mutation_type = qc_dict[k]['TYPE']
            expected_VAF = qc_dict[k]['TARGET_FREQUENCY']
            expected_hgvsc = qc_dict[k]['MUTATION_CDS']
            expected_hgvsp = qc_dict[k]['MUTATION_AA']
            mutation_id = qc_dict[k]['MUTATION_ID']
        
            selected_fields = [gene, ref, alt, mutation_type, expected_VAF, 
                               expected_hgvsc, expected_hgvsp, mutation_id]
            
            return selected_fields
        
        def process_final_qc_fields(qc_dict_fields, variant_summary_fields):

            def merge_two_dicts(x, y):
                """Given two dicts, merge them into a new dict as a shallow copy."""
                z = x.copy()
                z.update(y)
                return z
            
            other_fields = [comment, pass_or_fail]
            other_keys = ['comment', 'pass_or_fail']
            
            variant_summary_fields_keys = ['chrom', 'pos', 'status', 'main_consequence',
                                           't_vaf', 'n_vaf', 'transcript_link']
            
            qc_dict_fields_keys = ['gene', 'ref', 'alt', 'mutation_type', 'expected_VAF',
                                   'expected_hgvsc', 'expected_hgvsp', 'mutation_id']
            
            tmp_dict1 = dict(zip(variant_summary_fields_keys, variant_summary_fields))
            tmp_dict2 = dict(zip(qc_dict_fields_keys, qc_dict_fields))
            tmp_dict3 = dict(zip(other_keys, other_fields))
            
            merged_dict = merge_two_dicts(tmp_dict1, tmp_dict2)
            merged_dict = merge_two_dicts(merged_dict, tmp_dict3)
            
            output_fields = [
                             merged_dict['comment'],
                             merged_dict['pass_or_fail'],
                             merged_dict['chrom'],
                             merged_dict['gene'],
                             merged_dict['pos'],
                             merged_dict['ref'],
                             merged_dict['alt'],
                             merged_dict['mutation_type'],
                             merged_dict['status'],
                             merged_dict['main_consequence'],
                             merged_dict['t_vaf'],
                             merged_dict['n_vaf'],
                             merged_dict['expected_VAF'],
                             merged_dict['transcript_link'],
                             merged_dict['expected_hgvsc'],
                             merged_dict['expected_hgvsp'],
                             merged_dict['mutation_id']
                             ]
            
            #print merged_dict
            
            #print output_fields
            return output_fields
            
        for k1 in qc_dict.keys():
            if k1 in variant_summary_dict.keys():
                # Use k1 for key in both qc_dict and variant_summary_dict
                comment, pass_or_fail = expected_observed_VAF_test(k1, k1)
                summary_sheet[variant_summary_dict[k1]['comment_coordinate']] = comment
                # define fields to print from the two dictionaries
                selected_variant_dict_fields = select_final_qc_fields_from_variant_summary_dict(k1)
                selected_qc_dict_fields = select_final_qc_fields_from_qc_dict(k1)
                qc_sheet_output_fields = process_final_qc_fields(selected_qc_dict_fields, selected_variant_dict_fields)
            else:
                # Set match_bool = False.  If we encounter a gene with matching CDS or AA change, set to True and pass subsequent entries
                match_bool = False
                
                for k2 in variant_summary_dict.keys():
                    if qc_dict[k1]['GENE'] == variant_summary_dict[k2]['gene'] and match_bool is False:
                        if re.search(re.escape(str(qc_dict[k1]['MUTATION_CDS'])), str(variant_summary_dict[k2]['hgvsc']), re.IGNORECASE):
                            comment, pass_or_fail = expected_observed_VAF_test(k1, k2)
                            summary_sheet[variant_summary_dict[k2]['comment_coordinate']] = comment
                            selected_variant_dict_fields = select_final_qc_fields_from_variant_summary_dict(k2)
                            selected_qc_dict_fields = select_final_qc_fields_from_qc_dict(k1)
                            qc_sheet_output_fields = process_final_qc_fields(selected_qc_dict_fields, selected_variant_dict_fields)
                            match_bool = True                      
                        elif re.search(re.escape(str(qc_dict[k1]['MUTATION_AA'])), str(variant_summary_dict[k2]['hgvsp']), re.IGNORECASE):
                            comment, pass_or_fail = expected_observed_VAF_test(k1, k2)
                            summary_sheet[variant_summary_dict[k2]['comment_coordinate']] = comment
                            selected_variant_dict_fields = select_final_qc_fields_from_variant_summary_dict(k2)
                            selected_qc_dict_fields = select_final_qc_fields_from_qc_dict(k1)
                            qc_sheet_output_fields = process_final_qc_fields(selected_qc_dict_fields, selected_variant_dict_fields)
                            match_bool = True
                    elif match_bool is True:
                        pass
                    else:
                        comment = "NOT DETECTED"
                        pass_or_fail = "FAIL"
                        selected_variant_dict_fields = select_final_qc_fields_from_variant_summary_dict(k1)
                        selected_qc_dict_fields = select_final_qc_fields_from_qc_dict(k1)
                        qc_sheet_output_fields = process_final_qc_fields(selected_qc_dict_fields, selected_variant_dict_fields)

            # loop over columns and write them to the spreadsheet
            
            column_counter = 1
            for field in qc_sheet_output_fields:
                cell = qc_sheet.cell(column=column_counter, row=row_counter)
                cell.font = Font(name="Liberation Sans",size=10)
                cell.value = field
                if column_counter == 5 or column_counter == 14:
                    # Handle hyperlink columns differently
                    #
                    # Create IGV link 
                    # for position column
                    if column_counter == 5:
                        cell.value = '=HYPERLINK("http://localhost:60151/goto?locus=%s:%s", "%s")' % (qc_sheet_output_fields[2],
                                                                                                      qc_sheet_output_fields[4],
                                                                                                      qc_sheet_output_fields[4]
                                                                                                    )
                    if column_counter == 14:
                        print cell.value
                    cell.font = Font(name="Liberation Sans",size=10,color=BLUE)
                
                column_counter += 1
            
            row_counter += 1
            
    
    qc_dict = read_and_parse_QC_bed()
    summary_sheet = wb.get_sheet_by_name('summary')
    qc_sheet = create_qc_sheet(wb)

    row_counter = write_qc_header(qc_sheet)
    
    pp = pprint.PrettyPrinter(indent=4)
    
    write_qc_variants(variant_summary_dict, qc_dict, row_counter, qc_sheet, summary_sheet)

  
def main():
    parse_command_line(opts)
    try:
        wb = openpyxl.load_workbook(opts.input, data_only=True)
    except:
        print "ERROR: Could not open input file..."
        print "ERROR: Program is exiting..."
        sys.exit(1)
    
    try:
        sheet_name_and_column_slice_matrix = [
                                              ["transcripts","A1","R1"],
                                              ["co-located","A1","K1"],
                                              ["summary","A1","AX1"],
                                              ["cnv","A1","K1"],
                                              ["fusion","A1","L1"]
                                              ]
        transcript_id_row_dict = {}
        colocated_id_row_dict = {}
    except:
        print "ERROR: Could not initialize variables"
        print "ERROR: Program is exiting..."
        sys.exit(1)
    
    for sheet in sheet_name_and_column_slice_matrix:
        sheet = Sheet_name_and_column_slice_matrix(sheet)
        headers_index_dict = define_header_indices(wb,sheet.sheet_name,sheet.header_slice_start,sheet.header_slice_end)
        if sheet.sheet_name == "transcripts":
            indices = Transcript_headers_indices(headers_index_dict)
            transcript_id_row_dict = recreate_transcript_hyperlinks(wb,sheet.sheet_name, indices, transcript_id_row_dict)
        elif sheet.sheet_name == "co-located":
            indices = Colocated_headers_indices(headers_index_dict)
            colocated_id_row_dict = recreate_colocated_hyperlinks(wb,sheet.sheet_name, indices, colocated_id_row_dict)        
        elif sheet.sheet_name == "summary":
            indices = Summary_headers_indices(headers_index_dict)
            variant_summary_dict = apply_summary_filters(wb,sheet.sheet_name,indices,transcript_id_row_dict,colocated_id_row_dict)
        elif sheet.sheet_name == "cnv":
            indices = CNV_headers_indices(headers_index_dict)
            apply_CNV_filter(wb,sheet.sheet_name, indices)
        elif sheet.sheet_name == "fusion":
            if headers_index_dict == 0:
                pass
            else:
                indices = Fusion_headers_indices(headers_index_dict)
                apply_fusion_filter(wb,sheet.sheet_name, indices)
    
    if opts.enable_monthly_qc:
        perform_monthly_QC_check(wb, variant_summary_dict)
    
    try:
        wb.save(opts.output)
    except Exception, e:
        print "ERROR: Could not save output file..."
        print "ERROR: Program is exiting..."
        print str(e)
        sys.exit(1)

if __name__ == "__main__":
    main()
