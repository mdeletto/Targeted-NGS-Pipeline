#/usr/bin/python

import subprocess
import sys
import re
import os
import json
import datetime
import openpyxl
import errno
from pprint import PrettyPrinter
import pprint
from collections import defaultdict
import vcf
import shutil
import openpyxl
import string
from openpyxl.reader.excel import load_workbook
from subprocess import CalledProcessError


def mkdir_p(path):
    try:
        os.makedirs(path)
    except OSError as exc:  # Python >2.5
        if exc.errno == errno.EEXIST and os.path.isdir(path):
            pass
        else:
            raise

def samtools_mpileup(Samtools,tumor_bam,normal_bam,base_output,reference_fasta,regions_file):
    print "Processing samtools mpileup..."
    try:
        print "%s mpileup -Q 7 -h 50 -o 10 -e 17 -m 4 -f %s -l %s %s %s 2>> /tmp/error > %s.mpileup" % (Samtools,reference_fasta,regions_file,normal_bam,tumor_bam,base_output)
        subprocess.call("%s mpileup -Q 7 -h 50 -o 10 -e 17 -m 4 -f %s -l %s %s %s 2>> /tmp/error > %s.mpileup" % (Samtools,reference_fasta,regions_file,normal_bam,tumor_bam,base_output),shell=True) 
    except:
        print "ERROR: Samtools mpileup failed.  Aborting..."
        sys.exit()
    
def varscan_command(VarScan,mpileup,base_output):
    """Perform VarScan variant calling and apply various filtering/grouping - RETIRED"""
    print "Processing VarScan variant calling..."
    try:
        print "java -jar VarScan.jar somatic %s.mpileup %s --mpileup 1 --somatic-p-value 0.05 --min-coverage-tumor 20 --min-coverage-normal 5 --strand-filter 1 --output-vcf 1" % (mpileup,base_output)
        subprocess.call("java -jar %s somatic %s.mpileup %s --mpileup 1 --somatic-p-value 0.05 --min-coverage-tumor 20 --min-coverage-normal 5 --strand-filter 1 --output-vcf 1 2>> /tmp/error " % (VarScan,mpileup,base_output),shell=True)
        subprocess.call("java -jar %s processSomatic %s.indel.vcf 2>> /tmp/error" % (VarScan,base_output),shell=True)
        subprocess.call("java -jar %s processSomatic %s.snp.vcf 2>> /tmp/error" % (VarScan,base_output),shell=True)
        subprocess.call("vcf-concat %s.indel.vcf %s.snp.vcf 2>> /tmp/error > %s.snp.indel.vcf" % (base_output,base_output,base_output),shell=True)
        subprocess.call("vcf-sort -c %s.snp.indel.vcf 2>> /tmp/error > %s.snp.indel.sorted.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-concat %s*.hc.vcf 2>> /tmp/error > %s.snp.indel.hc.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-sort -c %s.snp.indel.hc.vcf 2>> /tmp/error > %s.snp.indel.sorted.hc.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-concat %s*Somatic.vcf 2>> /tmp/error > %s.snp.indel.somatic.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-sort -c %s.snp.indel.somatic.vcf 2>> /tmp/error > %s.snp.indel.sorted.somatic.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-concat %s*Germline.vcf 2>> /tmp/error > %s.snp.indel.germline.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-sort -c %s.snp.indel.germline.vcf 2>> /tmp/error > %s.snp.indel.sorted.germline.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-concat %s*LOH.vcf 2>> /tmp/error > %s.snp.indel.LOH.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-sort -c %s.snp.indel.LOH.vcf 2>> /tmp/error > %s.snp.indel.sorted.LOH.vcf" % (base_output,base_output),shell=True)
        subprocess.call("vcf-concat %s.snp.Somatic.vcf %s.indel.Somatic.vcf %s.snp.LOH.vcf %s.indel.LOH.vcf 2>> /tmp/error > %s.snp.indel.somatic_LOH.vcf" % (base_output,base_output,base_output,base_output,base_output),shell=True)
        subprocess.call("vcf-sort -c %s.snp.indel.somatic_LOH.vcf 2>> /tmp/error > %s.snp.indel.sorted.somatic_LOH.vcf" % (base_output,base_output),shell=True)
    except:
        print "ERROR: VarScan somatic variant calling failed.  Aborting..."
        sys.exit()
 
def SnpSift_filter(vcf_in, SnpSift, BEDTOOLS_EXE, regex_filter, base_output, program, do_not_separate_LOH_from_germline):
    """Apply filter via SnpSift and subtract LOH calls from the GERMLINE VCF --- no point in having duplicates in there...
    ...unless...well, you want them in there...
    """
    print "Filtering VCF input for %s..." % program
    try:
        if program == "ionreporter.germline" and do_not_separate_LOH_from_germline is False:
            subprocess.call("""awk 'BEGIN { FS = "\t" } ; { if ($10!="." && $11!=".") print}' %s | \
                               java -jar %s filter "%s" 2>> %s.snpsift.log | \
                               java -jar %s varType - | \
                               %s subtract -a stdin -b %s.ionreporter.loh.vcf > %s.%s.vcf""" % (vcf_in,SnpSift,regex_filter,base_output,SnpSift,BEDTOOLS_EXE,base_output,base_output,program),shell=True)
        else:
            subprocess.call("""awk 'BEGIN { FS = "\t" } ; { if ($10!="." && $11!=".") print}' %s | \
                                java -jar %s varType - | \
                               java -jar %s filter "%s" 2>> %s.snpsift.log > %s.%s.vcf""" % (vcf_in,SnpSift,SnpSift,regex_filter,base_output,base_output,program),shell=True)
        return "%s.%s.vcf" % (base_output, program)
    except:
        print "ERROR: Filtering of VCF input with SnpSift failed.  Aborting..."


def VEP_command_unfiltered(VEP,REF_FASTA,base_output,program, vcf_in):
    print "Annotating file..."
    try:
        subprocess.call('perl %s --quiet --cache --merged --offline --fasta %s -i %s --everything --check_alleles --coding_only --cache_version 83 --json -o %s.%s.json -fork 16 &> %s.vep.log' % (VEP,REF_FASTA,vcf_in,base_output,program,base_output),shell=True)
    except:
        print "ERROR: Could not initiate annotation on VCF file"


def VEP_command_filtered(VEP,REF_FASTA,base_output,program,vcf_in):
    try:
        def VEP_json_consequence_filtering(VEP_json_command_output,base_output,program):
            
            output_file = open("%s/%s.%s.json" % (os.getcwd(),base_output,program),"w")
            
            excluded_SO_terms_count = defaultdict(int)

            excluded_SO_terms = ['synonymous_variant',
                                 'start_retained_variant',
                                 'stop_retained_variant',
                                 'non_coding_transcript_exon_variant',
                                 'intron_variant',
                                 'downstream_gene_variant',
                                 'upstream_gene_variant',
                                 'UTR_variant',
                                 '3_prime_UTR_variant',
                                 '5_prime_UTR_variant',
                                 'NMD_transcript_variant',
                                 'null']            
            
            for line in VEP_json_command_output.strip().split("\n"):
                loaded = json.loads(line)
                if loaded['most_severe_consequence'] in excluded_SO_terms:
                    excluded_SO_terms_count[loaded['most_severe_consequence']] += 1
                else:
                    consequences_passing_filters = []
                    try:
                        for transcript_consequence in loaded['transcript_consequences']:
                            #pprint.pprint(transcript_consequence)
                            
                            for consequence_term in transcript_consequence['consequence_terms']:
                                if consequence_term in excluded_SO_terms:
                                    
                                    try:
                                        if transcript_consequence['canonical'] == 1:
                                            consequences_passing_filters.append(transcript_consequence)
                                        else:
                                            excluded_SO_terms_count[consequence_term] += 1
                                            pass
                                    except KeyError, e:
                                        pass
                                    except Exception, e:
                                        print "ERROR: Unknown exception - %s" % str(e)
    
                                else:
                                    consequences_passing_filters.append(transcript_consequence)
                            
                
                        
                        loaded['transcript_consequences'] = consequences_passing_filters            
                        
                        output_file.write(json.dumps(loaded)+"\n")
                    
                    except Exception, e:
                        print "ERROR: Unknown exception - %s" % str(e)
                        print "ERROR: Printing line with error:"
                        print loaded
                    
            output_file.close()
            
            print "Number of variants excluded from %s due to consequential filtering..." % program
            pp = pprint.PrettyPrinter(indent=4)
            pp.pprint(dict(excluded_SO_terms_count))

        with open("%s.vep.log" % base_output,"w") as err:
            vep_json_command = 'perl %s --quiet --cache --merged --offline --force_overwrite --fasta %s -i %s --everything --maf_exac --check_alleles --cache_version 83 --no_intergenic --minimal --allele_number --json --fork 16 --output_file STDOUT' % (VEP, REF_FASTA, vcf_in)
            vep_json_command_process = subprocess.Popen(vep_json_command,
                                                        stdout=subprocess.PIPE,
                                                        stderr=err,
                                                        shell=True)

        output, err = vep_json_command_process.communicate()
        
        
        VEP_json_consequence_filtering(output, base_output, program)
        
    except Exception,e:
        print "ERROR: VEP could not annotate and filter variants"
        print "ERROR: %s" % str(e)


def move_files_to_new_subdirectory(tumor_bam, normal_bam, base_output,galaxy_flag):
    print "Moving files to new directory named %s to prepare for final spreadsheet processing..." % base_output
    try:
        #subprocess.call('/usr/lib/jvm/java-8-oracle/jre/bin/java -jar %s -f %s.varscan.json %s.ionreporter.no_cnv.json %s.ionreporter.cnv.vcf %s.ionreporter.tsv' % (EDDY,base_output,base_output,base_output,base_output),shell=True)
        if os.getcwd().split("/")[-1] == base_output:

            if galaxy_flag is False:
                mkdir_p("%s/tmp" % os.getcwd())
                subprocess.call('mv %s*tmp* tmp/ 2>> /tmp/error' % base_output, shell=True)
                subprocess.call('mv %s*temp* tmp/ 2>> /tmp/error' % base_output, shell=True)           
    
                mkdir_p("%s/BAMs" % os.getcwd())
                if re.search("Population", normal_bam, re.IGNORECASE):
                    subprocess.call('mv %s* BAMs/ 2>> /tmp/error' % tumor_bam, shell=True)
                else:
                    subprocess.call('mv %s* %s* BAMs/ 2>> /tmp/error' % (tumor_bam, normal_bam),shell=True)
            
        else:
            
            mkdir_p("%s/%s" %(os.getcwd(),base_output))
            
            if galaxy_flag is False:
                mkdir_p("%s/%s/tmp" %(os.getcwd(),base_output))
                subprocess.call('mv %s*tmp* %s/tmp/ 2>> /tmp/error' % (base_output,base_output),shell=True)
                subprocess.call('mv %s*temp* %s/tmp/ 2>> /tmp/error' % (base_output,base_output),shell=True)           
    
                mkdir_p("%s/%s/BAMs" %(os.getcwd(),base_output))
                if re.search("Population", normal_bam, re.IGNORECASE):
                    subprocess.call('mv %s* %s/BAMs/ 2>> /tmp/error' % (tumor_bam, base_output),shell=True)
                else:
                    subprocess.call('mv %s* %s* %s/BAMs/ 2>> /tmp/error' % (tumor_bam, normal_bam, base_output),shell=True)
            subprocess.call('mv %s* %s/ 2>> /tmp/error' % (base_output,base_output),shell=True)
    except:
        print "ERROR: Failed to move files"
    try:
        subprocess.call('mv %s.ionreporter.fusions.vcf %s/%s.ionreporter.fusions.vcf 2>> /tmp/error' % (base_output,base_output,base_output),shell=True)
    except:
        print "ERROR: Could not move fusion file--it may not be defined."
 
def Galaxy_special_function(base_output,galaxy_output_directory):
    print "Processing moving files to Galaxy directory..."
    try:
        subprocess.call('mkdir %s' % galaxy_output_directory,shell=True)
        subprocess.call('cp -r %s/* %s' % (base_output,galaxy_output_directory),shell=True)
    except:
        print "ERROR: Could not move to Galaxy output directory..."

def edit_IR_tsv_file(ionreporter_version,ionreporter_tsv,base_output):
    if ionreporter_version==4.0:
        subprocess.call("cp %s %s.ionreporter.cnv.tsv" % (ionreporter_tsv,base_output),shell=True)
    elif ionreporter_version==4.4:
        tsv_output = open("%s.ionreporter.cnv.tsv" % base_output, "w")
        with open("%s" % ionreporter_tsv,"r") as f:
            for line in f.readlines():
                line = re.sub("# locus","#chr\tpos",line)
                match = re.search("chr(.{1,2}):(\d*)",line)
		if match is not None:
                    line = re.sub("chr(.{1,2}):(\d*)","chr%s\t%s" % (match.group(1),match.group(2)),line)
		tsv_output.write(line)
                    
    else:
        tsv_output = open("%s.ionreporter.tsv" % base_output, "w")
        with open("%s" % ionreporter_tsv,"r") as f:
            for line in f.readlines():
		line = re.sub("# locus","#chr\tpos",line)
                match = re.search("chr(.{1,2}):(\d*)",line)
                if match is not None:
                    line = re.sub("chr(.{1,2}):(\d*)","%s\t%s" % (match.group(1),match.group(2)),line)
                tsv_output.write(line)

def IR4_4_VCF_fix(vcf,base_output):
    vcf_fix = open("%s.ionreporter.somatic.unfiltered.vcf" % base_output,"w") 
    with open("%s" % vcf,"r") as f:
        for line in f.readlines():
            if re.match("^#",line):
                if re.search('Description="+.*"+',line):
                    vcf_fix.write(line)
                elif re.search('Description=.+>',line):
                    match = re.search('Description=(.+)>',line)
                    subgroup = match.group(1)
                    line = re.sub(subgroup,'%s%s%s' % ('"',subgroup,'"'),line)
                    vcf_fix.write(line)
                else:
                    vcf_fix.write(line)
            else:
                vcf_fix.write(line)
    vcf_fix.close()
    return "%s.ionreporter.somatic.unfiltered.vcf" % base_output

def pull_BAM_from_url(url,base_output,sample_type):
    import requests
    resp = requests.get(url,auth=('downstream','downstream'))
    output = open("%s.%s.bam" % (base_output,sample_type),"w")
    output.write(resp.content)
    return "%s.%s.bam" % (base_output,sample_type)

def zip_files(base_output,galaxy_dir):
    try:
        os.chdir(galaxy_dir)
        subprocess.call('zip -q %s.zip %s.ionreporter.cnv.vcf %s.ionreporter.somatic.json %s.ionreporter.loh.json %s.ionreporter.germline.json %s.ionreporter.tsv %s.mutect2.somatic.json %s.strelka.somatic.json %s.ionreporter.fusions.vcf %s.specimen.json' % (base_output,base_output,base_output,base_output,base_output,base_output,base_output,base_output,base_output,base_output),shell=True)
    except:
        print "ERROR: Could not create zip file for Downstream upload. Check to make sure all output files exist."


def IR_locate_variant_zip(basename,ionreporter_id):
    try:
        proc = subprocess.Popen(["""curl -k -H "Authorization:UmxyUXNPR3M1Q2RsbS9NYjBHQjBIaUxFTFA5RkJhRHBaMmlSSXZJTjBmUnNmQ0t1NkhOSUlrMStiNHFIQm16UjNKN2NYMzNOT2czcytqc2RveEhqK3BBSHhZNEhpNmRDVmtQaGRUZ1Z5ZXVXazJMTllQemIvV3A5c2NHOTNxRmY" "https://10.80.157.179/webservices_42/rest/api/analysis?format=json&name=%s&id=%s" 2> /dev/null""" % (basename,ionreporter_id)],shell=True,stdout=subprocess.PIPE)
        output, err = proc.communicate()
        #print output
        #print err
    except Exception, e:
        print "Unable to communicate with server.  Check Authorization key, server address, and your network connectivity."
        print str(e)
        sys.exit(1)
    try:
        output = output.strip("[").strip("]")
        #print output
        try:
            data = json.loads(output)
        except:
            print "Could not load json string"
        unfiltered_variants = data['data_links']['unfiltered_variants']
        filtered_variants = data['data_links']['filtered_variants']
        #print unfiltered_variants
        return unfiltered_variants
    except Exception, e:
        print "Unable to process IonReporter links.  Aborting..."
        print str(e)
        sys.exit(1)

def IR_locate_germline_variant_zip(basename,ionreporter_id):
    try:
        proc = subprocess.Popen(["""curl -k -H "Authorization:UmxyUXNPR3M1Q2RsbS9NYjBHQjBIaUxFTFA5RkJhRHBaMmlSSXZJTjBmUnNmQ0t1NkhOSUlrMStiNHFIQm16UjNKN2NYMzNOT2czcytqc2RveEhqK3BBSHhZNEhpNmRDVmtQaGRUZ1Z5ZXVXazJMTllQemIvV3A5c2NHOTNxRmY" "https://10.80.157.179/webservices_42/rest/api/analysis?format=json&name=%s&id=%s" 2> /dev/null""" % (basename,ionreporter_id)],shell=True,stdout=subprocess.PIPE)
        output, err = proc.communicate()
    except:
        print "Unable to communicate with server.  Check Authorization key, server address, and your network connectivity."
        sys.exit(1)
    try:
        output = output.strip("[").strip("]")

        try:
            data = json.loads(output)
        except:
            print "Could not load json string"
        unfiltered_variants = data['data_links']['unfiltered_variants']
        filtered_variants = data['data_links']['filtered_variants']
        control_barcode = data['samples']['CONTROL']
        sample_barcode = data['samples']['SAMPLE']
        
        ret_vals = [unfiltered_variants,
                    sample_barcode,
                    control_barcode]
        
        return ret_vals
        
    except Exception, e:
        print str(e)
        print "Unable to process IonReporter links.  Aborting..."
        sys.exit(1)

def IR_download_somatic_variant_zip(basename,variant_link,analysis_type):
#     try:
#         proc = subprocess.Popen(["""curl -k -H "Authorization:UmxyUXNPR3M1Q2RsbS9NYjBHQjBIaUxFTFA5RkJhRHBaMmlSSXZJTjBmUnNmQ0t1NkhOSUlrMStiNHFIQm16UjNKN2NYMzNOT2czcytqc2RveEhqK3BBSHhZNEhpNmRDVmtQaGRUZ1Z5ZXVXazJMTllQemIvV3A5c2NHOTNxRmY" "%s" 2> /dev/null -o IR.zip; \
#                                     unzip IR.zip && unzip %s.zip; \
#                                     cp ./Variants/*/*.vcf %s.ionreporter.%s_temp.vcf && cp ./Variants/*/*.tsv %s.ionreporter.%s_temp.tsv; \
#                                     rm -rf %s.zip QC Variants Workflow_Settings""" % (variant_link,basename,basename,analysis_type,basename,analysis_type,basename)],shell=True,stdout=subprocess.PIPE)
# 
#         files = ["%s.ionreporter.%s_temp.vcf" % (basename, analysis_type),
#                  "%s.ionreporter.%s_temp.tsv" % (basename, analysis_type)]
#         
#         return files
#         
#     except:
#         print "Unable to download and/or unzip IonReporter files.  Aborting..."
#         sys.exit(1)

    try:
        proc = subprocess.Popen(["""curl -k -H "Authorization:UmxyUXNPR3M1Q2RsbS9NYjBHQjBIaUxFTFA5RkJhRHBaMmlSSXZJTjBmUnNmQ0t1NkhOSUlrMStiNHFIQm16UjNKN2NYMzNOT2czcytqc2RveEhqK3BBSHhZNEhpNmRDVmtQaGRUZ1Z5ZXVXazJMTllQemIvV3A5c2NHOTNxRmY" "%s" 2> /dev/null -o IR_somatic.zip; \
                                  unzip -q IR_somatic.zip; \
                                  rm -rf IR_somatic.zip; \
                                  unzip -q %s*.zip; \
                                  cp ./Variants/*/*.vcf %s.ionreporter.%s_temp.vcf && cp ./Variants/*/*.tsv %s.ionreporter.%s_temp.tsv; \
                                  rm -rf IR_somatic.zip %s*.zip QC Variants Workflow_Settings VER*.log""" % (variant_link,basename,basename,analysis_type,basename,analysis_type,basename)], shell=True, stdout=subprocess.PIPE)
        output, err = proc.communicate()
        files = ["%s.ionreporter.%s_temp.vcf" % (basename, analysis_type),
                 "%s.ionreporter.%s_temp.tsv" % (basename, analysis_type)]
        
        return files
    
    except:
        print "Unable to download and/or unzip IonReporter files.  Aborting..."
        sys.exit(1)

def IR_download_fusion_zip(variant_link,basename):
    try:
        proc = subprocess.Popen(["""curl -k -H "Authorization:UmxyUXNPR3M1Q2RsbS9NYjBHQjBIaUxFTFA5RkJhRHBaMmlSSXZJTjBmUnNmQ0t1NkhOSUlrMStiNHFIQm16UjNKN2NYMzNOT2czcytqc2RveEhqK3BBSHhZNEhpNmRDVmtQaGRUZ1Z5ZXVXazJMTllQemIvV3A5c2NHOTNxRmY" "%s" 2> /dev/null -o IR_fusion.zip; \
                                    unzip -q IR_fusion.zip; \
                                    rm -rf IR_fusion.zip; \
                                    unzip -q %s*.zip; \
                                    cp ./Variants/*/*.vcf %s.ionreporter.fusions.vcf; \
                                    rm -rf IR_fusion.zip %s*.zip QC Variants Workflow_Settings VER*.log""" % (variant_link,basename,basename,basename)],shell=True,stdout=subprocess.PIPE)
    except Exception, e:
        print str(e)
        print "Unable to download and/or unzip IonReporter Fusion files.  Aborting..."
        sys.exit(1)



def IR_download_germline_variant_zip(VCFLIB_DIR, basename, variant_link, analysis_type, germline_sample_dict):
    """Downloads a germline variant zip from IR.  
    
    Because IR calls germline variants separately on each BAM and reports them as such, use VCFLIB to merge the VCFs together.
    """
    try:
        subprocess.call("""curl -k -H "Authorization:UmxyUXNPR3M1Q2RsbS9NYjBHQjBIaUxFTFA5RkJhRHBaMmlSSXZJTjBmUnNmQ0t1NkhOSUlrMStiNHFIQm16UjNKN2NYMzNOT2czcytqc2RveEhqK3BBSHhZNEhpNmRDVmtQaGRUZ1Z5ZXVXazJMTllQemIvV3A5c2NHOTNxRmY" "%s" 2> /dev/null -o IR_germline.zip; \
                           unzip -q IR_germline.zip; \
                           rm -rf IR_germline.zip; \
                           unzip -q %s*.zip; \
                           cp ./Variants/%s/%s.vcf %s.ionreporter.%s_temp.sample.vcf; \
                           cp ./Variants/%s/%s.vcf %s.ionreporter.%s_temp.control.vcf; \
                           rm -rf IR_germline.zip %s*.zip QC Variants Workflow_Settings VER*.log""" % (variant_link, basename, germline_sample_dict['sample'], germline_sample_dict['sample'], basename, analysis_type, germline_sample_dict['control'], germline_sample_dict['control'], basename, analysis_type, basename), shell=True)
#         subprocess.call("unzip -qq IR.zip && unzip -qq %s.zip" % basename, shell=True)
#         subprocess.call("cp ./Variants/%s/%s.vcf %s.ionreporter.%s_temp.sample.vcf" % (germline_sample_dict['sample'],germline_sample_dict['sample'],basename,analysis_type), shell=True)
#         subprocess.call("cp ./Variants/%s/%s.vcf %s.ionreporter.%s_temp.control.vcf" % (germline_sample_dict['control'],germline_sample_dict['control'],basename,analysis_type), shell=True)
#         subprocess.call("rm -rf %s.zip IR.zip QC Variants Workflow_Settings" % basename, shell=True)
        
        vcf_list = ['%s.ionreporter.%s_temp.sample.vcf' % (basename, analysis_type),
                    '%s.ionreporter.%s_temp.control.vcf' % (basename, analysis_type)]
        
        germline_unfiltered_vcf = combine_vcf(VCFLIB_DIR, vcf_list, basename, "ionreporter.germline.unfiltered")
        
        for vcf in vcf_list:
            subprocess.call("rm %s" % vcf, shell=True)
            
        return germline_unfiltered_vcf
    except Exception,e:
        print str(e)
        print "Unable to download and/or unzip IonReporter GERMLINE files.  Aborting..."
        sys.exit(1)

def rename_fusion_vcf(vcf,basename):
    try:
        print vcf
        subprocess.call("mv %s %s.ionreporter.fusions.vcf" % (vcf,basename),shell=True)
    except:
        print "Unable to rename IonReporter fusion.  Aborting..."
        sys.exit(1)
   
def bam_readcount_command(bam_readcount_exe,sample_type,base_output,reference_fasta,regions_file,bam_file):
    try:
        print "Producing bam readcounts for %s. This will be used for VarScan fpfilter command..." % sample_type
        subprocess.call("%s -f %s -l %s %s > %s.%s.readcounts.txt 2> /tmp/Bioinformatics_Pipeline_v1.3.bamreadcount.error.log" %(bam_readcount_exe,reference_fasta,regions_file,bam_file,base_output,sample_type),shell=True)
        readcount_filename = "%s.%s.readcounts.txt" %(base_output,sample_type)
        return readcount_filename
    except:
        print "Unable to produce bam readcounts for %s - %s." % (base_output,sample_type)
        sys.exit(1)
        
def select_target_regions(regions):
    try:
        print "Selecting target regions..."
        if regions=="CCP":
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/CCP/CCPHSMV2_052013.bed"
        elif regions=="OCP" or regions=="OCA":
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/OCP/AmpliSeq_OCP/OCP.20150630.designed.bed"
        elif regions=="BRCA":
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/BRCA1-2/BRCA1_2.20131001.designed.bed"
        elif regions=="CHPv2" or regions=="HSM":
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/CHPv2/CHP2.20131001.designed.bed"
        elif regions=="TSC":
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/TSC1-TSC2/TSC1_2.designed.bed"
        elif regions=="TFNA":
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/TFNA/Yale_Thyroid_DNA_WG_99191_167.1.20160607/WG_99191_167.1.20160607.designed.bed"
        else:
            REGIONS_FILE = "/home/michael/YNHH/Reference_Files/CCPHSMV2_052013.bed"
            print "WARNING: No bed file was selected.  Defaulting to using CCP regions to capture as much data as possible."
        print "BED file selected as %s" %(REGIONS_FILE)
        return REGIONS_FILE
    except:
        print "ERROR: Failed to select target regions"
        sys.exit(1)

def write_logfile(opts):
    try:
        logfile = open("%s.log.txt" % opts.base_output,"w")
        logfile.write("Parameters for analysis:\n\n")
        logfile.write("Date of analysis: %s\n" % datetime.datetime.now())
        for key,value in opts.__dict__.iteritems():
            logfile.write("%s\t%s\n" % (key,value))
        logfile.close()
    except:
        print "WARNING: Failed to open log file"
        
def delete_VEP_index():
    try:
        subprocess.call("rm -rf /home/michael/YNHH/Reference_Files/FASTA/hg19.VEP.fasta.index",shell=True)
    except:
        print "ERROR: Could not delete VEP index"

def select_varscan_vcf_subset(opts,base_output):
    try:
        print "Selecting VarScan vcf subset..."
        print "%s is selected for VCF output" % opts.status
        if opts.status=='All':
            vcf = '%s.snp.indel.sorted.vcf' % base_output
        elif opts.status=='All-HC':
            vcf = '%s.snp.indel.sorted.hc.vcf' % base_output
        elif opts.status=='Somatic':
            vcf = '%s.snp.indel.sorted.somatic.vcf' % base_output
        elif opts.status=='LOH':
            vcf = '%s.snp.indel.sorted.LOH.vcf' % base_output
        elif opts.status=='Germline':
            vcf = '%s.snp.indel.sorted.germline.vcf' % base_output
        elif opts.status=='Somatic_LOH':
            vcf = '%s.snp.indel.sorted.somatic_LOH.vcf' % base_output
        else:
            vcf = '%s.snp.indel.sorted.somatic_LOH.vcf' % base_output
        return vcf
    except:
        print "ERROR: Could not select VarScan vcf subset!!!"

def multibreak_vcf(VCFLIB_DIR,vcf,base_output,description):
    try:
        subprocess.call("%s/vcfbreakmulti %s > %s.%s.multibreak.vcf" % (VCFLIB_DIR,vcf,base_output,description),shell=True)
        return "%s.%s.multibreak.vcf" % (base_output, description)
    except:
        print "ERROR: multibreak vcf failed"

def combine_vcf(VCFLIB_DIR,vcf_list,base_output,description):
    try:
        vcfs = " ".join(vcf_list)
        #print '%s/vcfcombine %s > %s.%s.vcf' % (VCFLIB_DIR,vcfs,base_output,description)
        subprocess.call('%s/vcfcombine %s > %s.%s.vcf' % (VCFLIB_DIR,vcfs,base_output,description),shell=True)
        return "%s.%s.vcf" % (base_output, description)
    except:
        print "ERROR: Could not join VCFs"
 
def muTect_caller_command(MUTECT_EXE,REGIONS_FILE,MUTECT_V1_PON,dbsnp_vcf,cosmic_vcf,REF_FASTA,normal_bam,tumor_bam,base_output):
    try:
        subprocess.call("java -jar %s --analysis_type MuTect --reference_sequence %s -L %s -normal_panel %s -cosmic %s -dbsnp %s --input_file:normal %s --input_file:tumor %s -vcf %s.mutect.somatic.unfiltered.vcf -o %s.mutect.stats.tsv --enable_extended_output --enable_qscore_output --max_alt_allele_in_normal_fraction 0.1 --max_alt_alleles_in_normal_count 10 --gap_events_threshold 30 --pir_median_threshold 4 -dcov 1000 2>> /tmp/error" % (MUTECT_EXE,REF_FASTA,REGIONS_FILE,MUTECT_V1_PON,cosmic_vcf,dbsnp_vcf,normal_bam,tumor_bam,base_output,base_output),shell=True)
        return "%s.mutect.somatic.unfiltered.vcf" % base_output
    except:
        print "ERROR: MuTect failed"

def muTect2_caller_command(GATK_LATEST_EXE,REGIONS_FILE,MUTECT2_V1_PON,dbsnp_vcf,cosmic_vcf,REF_FASTA,normal_bam,tumor_bam,base_output):
    try:
        print """#--------------------MUTECT2 SOMATIC CALLING--------------------#"""
        if re.search("Population", normal_bam, re.IGNORECASE):
            print "RUNNING MUTECT2 IN TUMOR-ONLY MODE"
            #mutect2_command = "java -jar %s --analysis_type MuTect2 --reference_sequence %s -L %s --normal_panel %s --cosmic %s --dbsnp %s --input_file:tumor %s -o %s.mutect2.somatic.unfiltered.vcf --max_alt_allele_in_normal_fraction 0.1 -nct 8 --minPruning 10 --kmerSize 60" % (GATK_LATEST_EXE,REF_FASTA,REGIONS_FILE,MUTECT2_V1_PON,cosmic_vcf,dbsnp_vcf,tumor_bam,base_output)
            mutect2_command = "java -jar %s --analysis_type MuTect2 --reference_sequence %s -L %s --normal_panel %s --cosmic %s --dbsnp %s --input_file:tumor %s -o %s.mutect2.somatic.unfiltered.vcf -nct 8 --minPruning 10 --kmerSize 60" % (GATK_LATEST_EXE,REF_FASTA,REGIONS_FILE,MUTECT2_V1_PON,cosmic_vcf,dbsnp_vcf,tumor_bam,base_output)

        else:
            mutect2_command = "java -jar %s --analysis_type MuTect2 --reference_sequence %s -L %s --normal_panel %s --cosmic %s --dbsnp %s --input_file:normal %s --input_file:tumor %s -o %s.mutect2.somatic.unfiltered.vcf --max_alt_allele_in_normal_fraction 0.1 -nct 8 --minPruning 10 --kmerSize 60" % (GATK_LATEST_EXE,REF_FASTA,REGIONS_FILE,MUTECT2_V1_PON,cosmic_vcf,dbsnp_vcf,normal_bam,tumor_bam,base_output)

        print "MUTECT2 SOMATIC CALLING COMMAND:", mutect2_command
        log = open('%s.mutect2.log' % base_output, 'w')
        subprocess.call(mutect2_command,shell=True, stdout=log, stderr=log)
        log.close()

        return "%s.mutect2.somatic.unfiltered.vcf" % base_output
    except Exception, e:
        print "ERROR: MuTect2 failed"
        print "ERROR" + str(e)

def GATK_snp_caller_command(GATK_EXE,tumor_bam,normal_bam,base_output,REF_FASTA,dbsnp_vcf):
    try:
        subprocess.call("java -jar %s --analysis_type UnifiedGenotyper -I %s -I %s -o %s.gatk.snps.germline.raw.vcf -R %s -D %s -dcov 1000 -nt 16 2>> /tmp/error" % (GATK_EXE,tumor_bam,normal_bam,base_output,REF_FASTA,dbsnp_vcf),shell=True)
        return "%s.gatk.snps.germline.raw.vcf" % base_output
    except:
        print "ERROR: GATK Germline SNP calling failed"

def GATK_indel_caller_command(GATK_EXE,tumor_bam,normal_bam,base_output,REF_FASTA,dbsnp_vcf):
    try:
        subprocess.call("java -jar %s --analysis_type UnifiedGenotyper -glm INDEL -I %s -I %s -o %s.gatk.indels.germline.raw.vcf -R %s -D %s -dcov 1000 -nt 16 2>> /tmp/error" % (GATK_EXE,tumor_bam,normal_bam,base_output,REF_FASTA,dbsnp_vcf),shell=True)
        return "%s.gatk.indels.germline.raw.vcf" % base_output
    except:
        print "ERROR: GATK Germline INDEL calling failed"

def Strelka_somatic_variant_calling_command(STRELKA_EXE,normal_bam,tumor_bam,REF_FASTA,STRELKA_CONFIG,base_output):
    try:
        strelka_analysis_dir = '%s/strelka_analysis' % os.getcwd()
        strelka_command = "perl %s --normal=%s --tumor=%s --ref=%s --config=%s --output-dir=%s; make -j 16 -C %s &> %s.strelka.log" % (STRELKA_EXE,os.path.abspath(normal_bam),os.path.abspath(tumor_bam),REF_FASTA,STRELKA_CONFIG,strelka_analysis_dir,strelka_analysis_dir, base_output)
        print """#--------------------STRELKA SOMATIC CALLING--------------------#"""
        print "STRELKA SOMATIC CALLING COMMAND:", strelka_command
        
        log = open('%s.strelka.log' % base_output, 'w')
           
        subprocess.call("perl %s --normal=%s --tumor=%s --ref=%s --config=%s --output-dir=%s" % (STRELKA_EXE,os.path.abspath(normal_bam),os.path.abspath(tumor_bam),REF_FASTA,STRELKA_CONFIG,strelka_analysis_dir),shell=True, stdout=log, stderr=log)
        subprocess.call("make -j 16 -C %s" % (strelka_analysis_dir), shell=True, stdout=log, stderr=log) 
        log.close()
        
        vcf_list = ['%s/results/passed.somatic.indels.vcf' % strelka_analysis_dir,
                    '%s/results/passed.somatic.snvs.vcf' % strelka_analysis_dir]
        print vcf_list
        
        return vcf_list
    except:
        print "ERROR: Strelka somatic variant calling failed"

def extra_file_cleanup():
    subprocess.call("rm -rf Variants Workflow_Settings QC STDOUT_summary.html VER*.log strelka_analysis",shell=True)

def determine_num_variants_in_vcf(vcf):
    """This is a quick and dirty way to count variants in a VCF (does not account for multi-allelic sites)"""
    try:
        num_variants = subprocess.check_output("grep -vc '^#' %s" % vcf,shell=True)
        num_variants = num_variants.strip("\n")
    except CalledProcessError as e:
        if int(e.output.strip("\n")) > 1:
            raise
        else:
            num_variants = 0

    return num_variants

def sample_attribute_autodetection(basename, pipeline_version, panel):
    print """#--------------------SAMPLE ATTRIBUTE AUTODETECTION--------------------#"""
    
    def redefine_Downstream_panel_name(panel):
        """Redefines panel name to be recognized by Downstream."""
        if panel=='HSM' or panel=='CHPv2':
            panel = '50 gene panel'
        elif panel=='CCP' or panel=='409':
            panel = '409 gene panel'
        elif panel=='OCP' or panel=='OCA':
            panel = 'Oncomine panel'
        else:
            panel = panel
        
        return panel
    
    def check_if_tpl_spreadsheet_exists():
        """Checks if TPL spreadsheet is mounted and exists."""
        if os.path.isfile("/media/Tumor_Profiling_N_Drive/Tumor Profiling Lab/Tumor Profiling Documents/2017 TP Stats.xlsx"):
            return "/media/Tumor_Profiling_N_Drive/Tumor Profiling Lab/Tumor Profiling Documents/2017 TP Stats.xlsx"
        else:
            print "WARNING: Could not find the TP spreadsheet.  Attempting to troubleshoot the error..."
            if os.path.isdir("/media/Tumor_Profiling_N_Drive/"):
                sys.exit("ERROR: N Drive is successfully mounted.  Where did the sheet go?  Re-check the filename")
            else:
                sys.exit("ERROR: N Drive is not mounted.  Exiting...")
    
    
    def col2num(col):
        """Converts Excel column in letter format to column number."""
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num
    
    def append_entry_to_dict(sheet_obj, sheet_entries):
        """Append entry in spreadsheet to dict."""
        
        def get_max_column_and_row(sheet_obj):
            """Takes a spreadsheet object and returns a list with [max_column,max_row].
            Bases max_row and max_column calculations on first "None" that is encountered.
            """
        
            row_counter = 1
            
            for row in sheet_obj.iter_rows():
                if row_counter > 1:
                    # if column 1 is None, max_row = this row
                    if sheet_obj.cell(row=row_counter, column=1).value is not None:
                        max_row = sheet_obj.cell(row=row_counter, column=1).row
                    else:
                        break
                else:
                    for _cell in row:
                        if _cell.value is None:
                            break
                        else:
                            max_column = _cell.column
                row_counter += 1
    
            return [max_column, max_row]
        
        
        max_column, max_row = get_max_column_and_row(sheet_obj)    
        
        headers = []
        row_counter = 1
        for row in sheet_obj.iter_rows("A1:%s%s" % (max_column, max_row)):
            if row_counter > 1:
                values = []
                for cell in row:
                    values.append(cell.value)
                sheet_entries[sheet_name].append(dict(zip(headers,values)))
            else:
                for cell in row:
                    headers.append(cell.value)
                    if cell.value is None: # or cell.value == "":
                        break
            row_counter += 1
    
    def reformat_copath_id_format(test_string):
        """Reformats coPath ID if in X15-03344 format with 0 before numbers.
        Sometimes, the ID is entered into the spreadsheet without a preceding 0.
        """
        
        if re.search("-\d{5}", test_string):
            if re.search("-0\d{4}", test_string):
                "WARNING: Reformatting coPath ID"
                match = re.search("(\w+?)-0(\d{4})", test_string)
                id_part_one = match.group(1)
                id_part_two = match.group(2)
                test_string = "%s-%s" % (id_part_one, id_part_two)
                return [test_string, True]
            else:
                return [test_string, False]
        elif re.search("-\d{4}", test_string):
            match = re.search("(.+)-(\d){4}", test_string)
            return [test_string, False]
        else:
            "WARNING: %s is not in proper format.  Chances are we won't find any automatic annotations when searching." % test_string
    
    def write_specimen_json(basename, output_match_dict):
        """Write $basename.specimen.json."""
        
        try:
            with open('%s.specimen.json' % basename, 'w') as specimen_json_out:
                json.dump(output_match_dict, specimen_json_out)
            print "SUCCESS: %s.specimen.json was created" % basename       
        except:
            print "FAIL: Could not create %s.specimen.json" % basename
    
    def dict_clean(items):
        result = {}
        for key, value in items:
            if value is None:
                value = 'Unknown'
            result[key] = value
        return result
    
    def search_sheet_entries_for_id(copath_test, sheet_name, sheet_entries):
        """Search sheet_entries dict for query ID."""
         
        output_match = {}
        for sheet_name in sheet_entries.keys():
            for entry in sheet_entries[sheet_name]:
                print sheet_name
                # need these headers to populate some output in specimen.json
                # will try to use as much info from standard pipeline variables (e.g. panel, pipeline_version)
                necessary_headers = ['CoPath #',
                                     'laser/ manual MD',
                                     '% Malignant cells',
                                     'Tumor type',
                                     'Tumor Source']
                missing_headers = []
                
                    
                if all(x in entry.keys() for x in necessary_headers):
                    if entry['CoPath #'] == copath_test:
                        print "SUCCESS: All headers detected"
                        print "SUCCESS: We have a match in the TP spreadsheet"
                
                        try:
                            output_match['pipeline_version'] = pipeline_version
                        except:
                            print "FAIL: Could not define 'pipeline_version'"
                            output_match['pipeline_version'] = ''
                        try:
                            output_match['dissection'] = entry['laser/ manual MD']
                        except:
                            print "FAIL: Could not define 'dissection'"
                            output_match['dissection'] = ''
                        try:
                            output_match['malignant_cells'] = "{0:.0f}%".format(float(entry['% Malignant cells']) * 100)
                        except:
                            print "FAIL: Could not define 'malignant_cells'"
                            output_match['malignant_cells'] = ''
                        try:
                            output_match['tumor'] = entry['Tumor type'] + ", " + entry['Tumor Source']
                        except:
                            print "FAIL: Could not define 'tumor'"
                            output_match['tumor'] = ''
                        try:
                            output_match['normal'] = entry['Normal Source']
                        except:
                            print "FAIL: Could not define 'normal'"
                            output_match['normal'] = ''
                        try:
                            output_match['requested_by'] = entry['Requesting Physician']
                        except:
                            print "FAIL: Could not define 'requested_by'"
                            output_match['requested_by'] = ''
                        try:
                            output_match['panel'] = panel
                        except:
                            print "FAIL: Could not define 'requested_by'"
                            output_match['panel'] = ''
                        try:
                            output_match['panel_version'] = '1.0'
                        except:
                            print "FAIL: Could not define 'panel_version'"
                            output_match['panel_version'] = '1.0'
                else:
                    for necessary_header in necessary_headers:
                        if not necessary_header in entry.keys():
                            if necessary_header != 'CoPath#' and 'CoPath #' in entry.keys():
                                # Let's slowly try to parse out certain values
                                try:
                                    output_match['pipeline_version'] = pipeline_version
                                except:
                                    print "FAIL: Could not define 'pipeline_version'"
                                    output_match['pipeline_version'] = ''
                                try:
                                    output_match['dissection'] = entry['laser/ manual MD']
                                except:
                                    print "FAIL: Could not define 'dissection'"
                                    output_match['dissection'] = ''
                                try:
                                    output_match['malignant_cells'] = "{0:.0f}%".format(float(entry['% Malignant cells']) * 100)
                                except:
                                    print "FAIL: Could not define 'malignant_cells'"
                                    output_match['malignant_cells'] = ''
                                try:
                                    output_match['tumor'] = entry['Tumor type'] + ", " + entry['Tumor Source']
                                except:
                                    print "FAIL: Could not define 'tumor'"
                                    output_match['tumor'] = ''
                                try:
                                    output_match['normal'] = entry['Normal Source']
                                except:
                                    print "FAIL: Could not define 'normal'"
                                    output_match['normal'] = ''
                                try:
                                    output_match['requested_by'] = entry['Requesting Physician']
                                except:
                                    print "FAIL: Could not define 'requested_by'"
                                    output_match['requested_by'] = ''
                                try:
                                    output_match['panel'] = panel
                                except:
                                    print "FAIL: Could not define 'requested_by'"
                                    output_match['panel'] = ''
                                try:
                                    output_match['panel_version'] = '1.0'
                                except:
                                    print "FAIL: Could not define 'panel_version'"
                                    output_match['panel_version'] = '1.0'
                                    
                            else:
                                # Create default specimen.json file if CoPath # does not exist in spreadsheet
                                print "FAIL: Need CoPath # to retrieve sample information.  Creating default specimen.json file"
                                
                                try:
                                    output_match['pipeline_version'] = pipeline_version
                                except:
                                    print "FAIL: Could not define 'pipeline_version'"
                                try:
                                    output_match['dissection'] = ""
                                except:
                                    print "FAIL: Could not define 'dissection'"
                                try:
                                    output_match['malignant_cells'] = ""
                                except:
                                    print "FAIL: Could not define 'malignant_cells'"
                                try:
                                    output_match['tumor'] = ""
                                except:
                                    print "FAIL: Could not define 'tumor'"
                                try:
                                    output_match['normal'] = ""
                                except:
                                    print "FAIL: Could not define 'normal'"
                                try:
                                    output_match['panel'] = panel
                                except:
                                    print "FAIL: Could not define 'requested_by'"
                                try:
                                    output_match['panel_version'] = '1.0'
                                except:
                                    print "FAIL: Could not define 'panel_version'"
                                    output_match['panel_version'] = '1.0'
    
        return output_match
    
    panel = redefine_Downstream_panel_name(panel)
    # Check if TP stats spreadsheet exists
    spreadsheet_path = check_if_tpl_spreadsheet_exists()
    # Load workbook
    spreadsheet = load_workbook(spreadsheet_path, data_only=True)
    # Get sheet names
    sheet_names = spreadsheet.get_sheet_names()
    # Initialize defaultdict
    sheet_entries = defaultdict(list)
    # Pop sheet_names that are incompatible with pipeline
    sheet_names.remove("TaqMan Cases")
    # Create a dict of entries from each sheet
    for sheet_name in sheet_names:
        sheet_obj = spreadsheet.get_sheet_by_name(sheet_name)
        append_entry_to_dict(sheet_obj, sheet_entries)


    # Search dict entries for ID that matches query ID
    output_match = search_sheet_entries_for_id(basename, sheet_name, sheet_entries)
    
    if not output_match:
        reformatted_basename, reformatted_bool = reformat_copath_id_format(basename)
        if reformatted_bool is True:
            output_match = search_sheet_entries_for_id(reformatted_basename, sheet_name, sheet_entries)
        else:
            print "WARNING: No matches on %s" % copath_test
    
    if output_match:
        dict_str = json.dumps(output_match)
        output_match = json.loads(dict_str, object_pairs_hook=dict_clean)
        pp = PrettyPrinter(indent=4)
        pp.pprint(output_match)
        write_specimen_json(basename, output_match)

def extract_fusion_VCF_information(vcf):
    """Extracts information from the ionreporter.fusions.vcf file."""
    
    fusion_dict = defaultdict(dict)
    fusion_dict['sum_control_count'] = 0
    
    with open(vcf) as f:
        for line in f.readlines():
            line = line.strip()
            
            if re.search("##mapd", line):
                match = re.search("##mapd=(.*)", line)
                mapd = match.group(1)
                fusion_dict['mapd'] = mapd
            elif re.search("SVTYPE=GeneExpression", line):
                split_line = line.split("\t")
                gene_match = re.search("GENE_NAME=(.+?);", line)
                gene = gene_match.group(1)
                read_count_match = re.search("READ_COUNT=(.+?);", line)
                read_count = read_count_match.group(1)
                fusion_dict['gene_expression_read_counts'][gene] = read_count
            elif re.search("SVTYPE=ExprControl", line):
                gene_match = re.search("GENE_NAME=(.+?);", line)
                gene = gene_match.group(1)
                read_count_match = re.search("READ_COUNT=(.+?);", line)
                read_count = read_count_match.group(1)
                fusion_dict['control_per_gene_read_counts'][gene] = read_count
                fusion_dict['sum_control_count'] += int(read_count)
            elif re.search("##TotalMappedFusionPanelReads", line):
                match = re.search("##TotalMappedFusionPanelReads=(.*)", line)
                total_mapped_fusion_reads = match.group(1)
                fusion_dict['total_mapped_fusion_reads'] = total_mapped_fusion_reads

    return fusion_dict